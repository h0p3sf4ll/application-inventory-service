from __future__ import annotations

import os
import time
import unittest
import uuid
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from appsec_scan_router.models import ScanConfig
from appsec_scan_router.postgres import (
    PostgresInventoryWriter,
    export_inventory_xlsx,
    search_inventory,
)

try:
    import psycopg
    from psycopg import sql
except ImportError:
    psycopg = None
    sql = None


POSTGRES_TEST_DSN = os.getenv("APPLICATION_INVENTORY_TEST_POSTGRES_DSN", "")


@unittest.skipUnless(
    POSTGRES_TEST_DSN and psycopg and sql,
    "PostgreSQL integration DSN is not configured",
)
class PostgresInventoryIntegrationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.schema = f"inventory_test_{uuid.uuid4().hex[:12]}"

    def tearDown(self) -> None:
        with psycopg.connect(POSTGRES_TEST_DSN, autocommit=True) as connection:
            connection.execute(
                sql.SQL("DROP SCHEMA IF EXISTS {schema} CASCADE").format(
                    schema=sql.Identifier(self.schema)
                )
            )

    def test_repeated_results_update_current_state_without_duplicates(self) -> None:
        self.write_result("user-a", "alice", "1.0.0")
        self.write_result("user-a", "alice", "1.0.0")
        self.write_result("user-a", "alice", "1.1.0")

        with psycopg.connect(POSTGRES_TEST_DSN) as connection:
            counts = connection.execute(
                sql.SQL(
                    """
                    SELECT
                        (SELECT count(*) FROM {flat_table}),
                        (SELECT count(*) FROM {repositories}),
                        (SELECT count(*) FROM {branch_inventory}),
                        (SELECT count(*) FROM {scan_runs}),
                        (SELECT count(*) FROM {inventory_types}),
                        (SELECT count(*) FROM {contributors}),
                        (SELECT count(*) FROM {web_domains}),
                        (SELECT count(*) FROM {web_domain_sources}),
                        (SELECT count(*) FROM {store_listings})
                    """
                ).format(
                    flat_table=sql.Identifier(
                        self.schema, "application_inventory_assets"
                    ),
                    repositories=sql.Identifier(self.schema, "repositories"),
                    branch_inventory=sql.Identifier(self.schema, "branch_inventory"),
                    scan_runs=sql.Identifier(self.schema, "scan_runs"),
                    inventory_types=sql.Identifier(self.schema, "inventory_types"),
                    contributors=sql.Identifier(self.schema, "branch_contributors"),
                    web_domains=sql.Identifier(self.schema, "web_domains"),
                    web_domain_sources=sql.Identifier(
                        self.schema, "web_domain_sources"
                    ),
                    store_listings=sql.Identifier(self.schema, "store_listings"),
                )
            ).fetchone()

        self.assertEqual(counts, (1, 1, 1, 1, 2, 2, 2, 2, 2))
        search = search_inventory(
            POSTGRES_TEST_DSN,
            schema=self.schema,
            owner_user_id="user-a",
            query="inventory api_service",
        )
        self.assertEqual(search["total"], 1)
        self.assertEqual(search["rows"][0]["inventory_version"], "1.1.0")
        self.assertEqual(
            search["rows"][0]["primary_web_domain"], "inventory.example.engineering"
        )
        self.assertEqual(
            search_inventory(
                POSTGRES_TEST_DSN,
                schema=self.schema,
                owner_user_id="user-a",
                query="inventory.example.engineering",
            )["total"],
            1,
        )
        structured = search_inventory(
            POSTGRES_TEST_DSN,
            schema=self.schema,
            owner_user_id="user-a",
            filters={
                "application_types": ["api_service"],
                "has_domain": True,
                "confidences": ["high"],
                "updated_within_days": 90,
            },
        )
        self.assertEqual(structured["total"], 1)
        self.assertEqual(structured["filters"]["application_types"], ["api_service"])
        column_filtered = search_inventory(
            POSTGRES_TEST_DSN,
            schema=self.schema,
            owner_user_id="user-a",
            filters={
                "application_search": "Inventory",
                "repository_search": "inventory-service",
                "branch_search": "main",
                "domain_search": "example.engineering",
                "sort_by": "domain",
                "sort_direction": "asc",
            },
        )
        self.assertEqual(column_filtered["total"], 1)
        self.assertEqual(column_filtered["filters"]["sort_by"], "domain")
        self.assertEqual(
            search_inventory(
                POSTGRES_TEST_DSN,
                schema=self.schema,
                owner_user_id="user-a",
                filters={"application_search": "Inventory%_"},
            )["total"],
            0,
        )

        export = export_inventory_xlsx(
            POSTGRES_TEST_DSN,
            schema=self.schema,
            owner_user_id="user-a",
            filters={"application_types": ["api_service"]},
        )
        workbook = load_workbook(BytesIO(export), read_only=True)
        self.assertEqual(sum(1 for _ in workbook["Inventory"].iter_rows()), 2)

        self.write_result("user-b", "bob", "2.0.0")

        with psycopg.connect(POSTGRES_TEST_DSN) as connection:
            counts = connection.execute(
                sql.SQL(
                    "SELECT (SELECT count(*) FROM {flat_table}), (SELECT count(*) FROM {repositories}), "
                    "(SELECT count(*) FROM {branch_inventory})"
                ).format(
                    flat_table=sql.Identifier(
                        self.schema, "application_inventory_assets"
                    ),
                    repositories=sql.Identifier(self.schema, "repositories"),
                    branch_inventory=sql.Identifier(self.schema, "branch_inventory"),
                )
            ).fetchone()

        self.assertEqual(counts, (2, 2, 2))
        self.assertEqual(
            search_inventory(
                POSTGRES_TEST_DSN, schema=self.schema, owner_user_id="user-a"
            )["total"],
            1,
        )

    def test_pending_rows_are_committed_for_live_search(self) -> None:
        config = self.scan_config("user-live", "live.user")
        result = self.sample_result("1.0.0")

        with patch.dict(
            os.environ,
            {
                "APPLICATION_INVENTORY_POSTGRES_COMMIT_ROWS": "1000",
                "APPLICATION_INVENTORY_POSTGRES_COMMIT_SECONDS": "0.05",
            },
        ):
            writer = PostgresInventoryWriter(config)
            writer.__enter__()
            try:
                writer.write_result(result)
                deadline = time.monotonic() + 2
                total = 0
                while time.monotonic() < deadline:
                    total = search_inventory(
                        POSTGRES_TEST_DSN,
                        schema=self.schema,
                        owner_user_id="user-live",
                    )["total"]
                    if total:
                        break
                    time.sleep(0.03)
                self.assertEqual(total, 1)
            finally:
                writer.close()

    def write_result(
        self, owner_user_id: str, owner_user_login: str, version: str
    ) -> None:
        with PostgresInventoryWriter(
            self.scan_config(owner_user_id, owner_user_login)
        ) as writer:
            writer.write_result(self.sample_result(version))

    def scan_config(self, owner_user_id: str, owner_user_login: str) -> ScanConfig:
        return ScanConfig(
            org="ExampleEngineering",
            pat="test-token",
            project=None,
            out_dir=Path("reports"),
            out_prefix="integration",
            max_workers=1,
            content_workers=1,
            max_commits_per_repo=0,
            timeout_seconds=30,
            min_confidence="medium",
            postgres_dsn=POSTGRES_TEST_DSN,
            postgres_schema=self.schema,
            owner_user_id=owner_user_id,
            owner_user_login=owner_user_login,
        )

    def sample_result(self, version: str) -> dict[str, object]:
        return {
            "provider": "azure-devops",
            "organization": "ExampleEngineering",
            "project": "Inventory",
            "repo_name": "inventory-service",
            "branch_name": "main",
            "inventory_name": "Inventory Service",
            "inventory_version": version,
            "inventory_types": "api_service; microservice",
            "categories": "python; fastapi",
            "branch_contributing_developers": "Alice <alice@example.com>; Bob <bob@example.com>",
            "confidence": "high",
            "score": 12,
            "last_updated": "2026-07-13T12:00:00+00:00",
            "branch_last_updated": "2026-07-13T12:00:00+00:00",
            "primary_web_domain": "inventory.example.engineering",
            "web_domains": "inventory.example.engineering; api.example.engineering",
            "web_urls": "https://inventory.example.engineering; https://api.example.engineering",
            "web_domain_status": "confirmed",
            "web_domain_sources": "inventory.example.engineering [github:deployment_status]",
            "web_domain_evidence": """[
                {
                    "domain": "inventory.example.engineering",
                    "url": "https://inventory.example.engineering",
                    "confidence": "confirmed",
                    "environment": "production",
                    "sources": ["github:deployment_status"]
                },
                {
                    "domain": "api.example.engineering",
                    "url": "https://api.example.engineering",
                    "confidence": "configured",
                    "environment": "production",
                    "sources": ["source:/deploy/ingress.yaml:host:3"]
                }
            ]""",
            "apple_app_store_name": "Inventory Mobile",
            "apple_app_store_identifier": "com.example.inventory",
            "apple_app_store_url": "https://apps.apple.com/app/id123456789",
            "apple_app_store_version": version,
            "apple_app_store_validation_passed": "TRUE",
            "apple_app_store_lookup_status": "found",
            "google_play_name": "Inventory Mobile",
            "google_play_identifier": "com.example.inventory",
            "google_play_url": "https://play.google.com/store/apps/details?id=com.example.inventory",
            "google_play_version": version,
            "google_play_validation_passed": "TRUE",
            "google_play_lookup_status": "found",
        }
