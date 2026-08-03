from __future__ import annotations

import os
import unittest
import uuid
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook

from appsec_scan_router.aspm_ingest import parse_finding_document
from appsec_scan_router.aspm_models import FindingDocument, FindingInput, SourceLocation
from appsec_scan_router.aspm_postgres import AspmRepository
from appsec_scan_router.postgres import create_database_schema

try:
    import psycopg
    from psycopg import sql
except ImportError:
    psycopg = None
    sql = None


POSTGRES_TEST_DSN = os.getenv("APPLICATION_INVENTORY_TEST_POSTGRES_DSN", "")


@unittest.skipUnless(
    POSTGRES_TEST_DSN and psycopg and sql,
    "PostgreSQL integration DSN is not configured",
)
class AspmPostgresIntegrationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.schema = f"aspm_test_{uuid.uuid4().hex[:12]}"
        with psycopg.connect(POSTGRES_TEST_DSN, autocommit=True) as connection:
            create_database_schema(
                connection, self.schema, "application_inventory_assets"
            )
            connection.execute(
                sql.SQL(
                    "INSERT INTO {runs} (scan_id, provider, organization, owner_user_id, owner_user_login, started_at) VALUES ('scan-1', 'github-enterprise', 'ExampleEngineering', 'user-a', 'alice', now())"
                ).format(runs=sql.Identifier(self.schema, "scan_runs"))
            )
            repository_id = connection.execute(
                sql.SQL(
                    "INSERT INTO {repositories} (owner_user_id, owner_user_login, provider, organization, project, repo_name, web_url) VALUES ('user-a', 'alice', 'github-enterprise', 'ExampleEngineering', '', 'payments-api', 'https://github.com/ExampleEngineering/payments-api') RETURNING repository_id"
                ).format(repositories=sql.Identifier(self.schema, "repositories"))
            ).fetchone()[0]
            self.branch_inventory_id = connection.execute(
                sql.SQL(
                    """
                    INSERT INTO {branches} (
                        repository_id, scan_id, owner_user_id, owner_user_login,
                        branch_name, inventory_name, primary_language, last_updated,
                        confidence, score, row_data, detection_evidence, scan_started_at
                    ) VALUES (%s, 'scan-1', 'user-a', 'alice', 'main', 'Payments API',
                              'Python', now(), 'high', 95, '{{}}'::jsonb, '{{}}'::jsonb, now())
                    RETURNING branch_inventory_id
                    """
                ).format(branches=sql.Identifier(self.schema, "branch_inventory")),
                (repository_id,),
            ).fetchone()[0]
            connection.execute(
                sql.SQL(
                    "INSERT INTO {types} (branch_inventory_id, inventory_type) VALUES (%s, 'api_service')"
                ).format(types=sql.Identifier(self.schema, "inventory_types")),
                (self.branch_inventory_id,),
            )
            connection.execute(
                sql.SQL(
                    "INSERT INTO {domains} (branch_inventory_id, domain, url, confidence, is_primary) VALUES (%s, 'payments.example.test', 'https://payments.example.test', 'confirmed', true)"
                ).format(domains=sql.Identifier(self.schema, "web_domains")),
                (self.branch_inventory_id,),
            )
        self.repository = AspmRepository(POSTGRES_TEST_DSN, self.schema)

    def tearDown(self) -> None:
        with psycopg.connect(POSTGRES_TEST_DSN, autocommit=True) as connection:
            connection.execute(
                sql.SQL("DROP SCHEMA IF EXISTS {schema} CASCADE").format(
                    schema=sql.Identifier(self.schema)
                )
            )

    def test_findings_are_linked_deduplicated_scored_and_reconciled(self) -> None:
        payload = {
            "format": "generic",
            "tool": {"key": "codeql", "name": "CodeQL", "type": "sast"},
            "context": {
                "provider": "github-enterprise",
                "organization": "ExampleEngineering",
                "repository": "payments-api",
                "branch": "main",
            },
            "findings": [
                {
                    "id": "finding-1",
                    "title": "SQL injection",
                    "severity": "high",
                    "rule_id": "py/sql-injection",
                    "path": "src/db.py",
                    "line": 20,
                    "cwe": "CWE-89",
                    "cvss_score": 8.2,
                }
            ],
            "completeSnapshot": True,
            "scannedTargets": [
                {
                    "provider": "github-enterprise",
                    "organization": "ExampleEngineering",
                    "repository": "payments-api",
                    "branch": "main",
                }
            ],
        }

        first = self.repository.ingest(
            "user-a", "alice", parse_finding_document(payload)
        )
        second = self.repository.ingest(
            "user-a", "alice", parse_finding_document(payload)
        )

        self.assertEqual(first["inserted"], 1)
        self.assertEqual(first["assetsCovered"], 1)
        self.assertEqual(second["inserted"], 0)
        self.assertEqual(second["updated"], 1)
        search = self.repository.search_findings("user-a")
        self.assertEqual(search["total"], 1)
        finding = search["rows"][0]
        self.assertEqual(finding["branch_inventory_id"], self.branch_inventory_id)
        self.assertEqual(finding["primary_web_domain"], "payments.example.test")
        self.assertGreaterEqual(finding["risk_score"], 60)
        self.assertEqual(search["facets"]["tools"][0]["value"], "codeql")
        self.assertEqual(search["facets"]["tools"][0]["label"], "CodeQL")
        posture = self.repository.posture("user-a")
        self.assertEqual(posture["summary"]["assets"], 1)
        self.assertEqual(posture["summary"]["active_findings"], 1)
        self.assertEqual(posture["coverage"]["coverage_percent"], 100.0)
        with psycopg.connect(POSTGRES_TEST_DSN, autocommit=True) as connection:
            repository_id = connection.execute(
                sql.SQL(
                    "SELECT repository_id FROM {branches} WHERE branch_inventory_id = %s"
                ).format(branches=sql.Identifier(self.schema, "branch_inventory")),
                (self.branch_inventory_id,),
            ).fetchone()[0]
            connection.execute(
                sql.SQL(
                    """
                    INSERT INTO {branches} (
                        repository_id, scan_id, owner_user_id, owner_user_login,
                        branch_name, inventory_name, primary_language, last_updated,
                        confidence, score, row_data, detection_evidence, scan_started_at
                    ) VALUES (%s, 'scan-1', 'user-a', 'alice', 'inactive', 'Inactive API',
                              'Python', now(), 'high', 95, '{{}}'::jsonb, '{{}}'::jsonb, now())
                    """
                ).format(branches=sql.Identifier(self.schema, "branch_inventory")),
                (repository_id,),
            )
        posture = self.repository.posture("user-a")
        self.assertEqual(posture["summary"]["assets"], 2)
        self.assertEqual(
            posture["summary"]["average_risk"],
            posture["topAssets"][0]["max_risk_score"],
        )
        coverage = self.repository.coverage("user-a")
        active_coverage = next(
            row
            for row in coverage["rows"]
            if row["branch_inventory_id"] == self.branch_inventory_id
        )
        self.assertEqual(active_coverage["coverage_status"], "current")
        self.assertEqual(active_coverage["tools"], "CodeQL")

        updated = self.repository.update_finding(
            "user-a",
            "alice",
            finding["finding_id"],
            "triaged",
            assignee="payments-team",
            note="Confirmed by AppSec.",
        )
        self.assertEqual(updated["status"], "triaged")
        detail = self.repository.finding_detail("user-a", finding["finding_id"])
        self.assertEqual(detail["finding"]["assignee"], "payments-team")
        self.assertEqual(detail["events"][0]["event_type"], "workflow_updated")

        empty_snapshot = parse_finding_document(
            {
                "format": "generic",
                "tool": {"key": "codeql", "name": "CodeQL", "type": "sast"},
                "findings": [],
                "completeSnapshot": True,
                "scannedTargets": payload["scannedTargets"],
            }
        )
        reconciled = self.repository.ingest("user-a", "alice", empty_snapshot)
        self.assertEqual(reconciled["resolved"], 1)
        resolved = self.repository.search_findings(
            "user-a", filters={"statuses": ["resolved"]}
        )
        self.assertEqual(resolved["total"], 1)

        workbook = load_workbook(
            BytesIO(self.repository.export_findings("user-a", "xlsx")),
            read_only=True,
        )
        self.assertIn("Security Findings", workbook.sheetnames)
        self.assertEqual(sum(1 for _ in workbook["Security Findings"].iter_rows()), 2)

    def test_asset_profile_changes_recalculate_risk_and_owner_scope_is_enforced(
        self,
    ) -> None:
        payload = {
            "format": "generic",
            "tool": {"key": "sca", "name": "Dependency Scanner", "type": "sca"},
            "context": {
                "organization": "ExampleEngineering",
                "repository": "payments-api",
                "branch": "main",
            },
            "findings": [
                {
                    "id": "dependency-1",
                    "title": "Outdated dependency",
                    "severity": "medium",
                }
            ],
        }
        self.repository.ingest("user-a", "alice", parse_finding_document(payload))
        before = self.repository.search_findings("user-a")["rows"][0]["risk_score"]
        profile = self.repository.update_asset_profile(
            "user-a",
            "alice",
            self.branch_inventory_id,
            {
                "criticality": "mission_critical",
                "internetExposed": True,
                "dataClassification": "restricted",
                "businessOwner": "Payments",
                "technicalOwner": "payments-team",
                "tags": ["pci", "tier-0"],
            },
        )
        after = self.repository.search_findings("user-a")["rows"][0]["risk_score"]

        self.assertEqual(profile["criticality"], "mission_critical")
        stored_profile = self.repository.asset_profile(
            "user-a", self.branch_inventory_id
        )
        self.assertEqual(stored_profile["technical_owner"], "payments-team")
        self.assertEqual(stored_profile["tags"], ["pci", "tier-0"])
        self.assertGreater(after, before)
        self.assertEqual(self.repository.search_findings("user-b")["total"], 0)
        with self.assertRaises(KeyError):
            self.repository.update_asset_profile(
                "user-b", "bob", self.branch_inventory_id, {}
            )

    def test_failed_import_is_auditable_and_atomic(self) -> None:
        document = parse_finding_document(
            {
                "format": "generic",
                "tool": {"key": "failing-tool", "name": "Failing Tool"},
                "findings": [
                    {
                        "id": "finding-1",
                        "title": "Finding that cannot be persisted",
                        "severity": "high",
                    }
                ],
            }
        )

        with patch.object(
            self.repository,
            "_upsert_finding",
            side_effect=RuntimeError("simulated persistence failure"),
        ):
            with self.assertRaisesRegex(RuntimeError, "simulated persistence failure"):
                self.repository.ingest("user-a", "alice", document)

        with psycopg.connect(
            POSTGRES_TEST_DSN, row_factory=psycopg.rows.dict_row
        ) as connection:
            imported = connection.execute(
                sql.SQL(
                    "SELECT status, error_count, error_message FROM {imports} WHERE owner_user_id = 'user-a' ORDER BY started_at DESC LIMIT 1"
                ).format(imports=sql.Identifier(self.schema, "aspm_imports"))
            ).fetchone()
            finding_count = connection.execute(
                sql.SQL("SELECT count(*) AS finding_count FROM {findings}").format(
                    findings=sql.Identifier(self.schema, "aspm_findings")
                )
            ).fetchone()["finding_count"]

        self.assertEqual(imported["status"], "failed")
        self.assertEqual(imported["error_count"], 1)
        self.assertIn("simulated persistence failure", imported["error_message"])
        self.assertEqual(finding_count, 0)

    def test_posture_reconciles_abandoned_import_and_connector_sync(self) -> None:
        self.repository.ensure_schema()
        with psycopg.connect(POSTGRES_TEST_DSN, autocommit=True) as connection:
            tool_id = connection.execute(
                sql.SQL(
                    "INSERT INTO {tools} (owner_user_id, tool_key, tool_name, tool_type) "
                    "VALUES ('user-a', 'semgrep', 'Semgrep', 'sast') RETURNING tool_id"
                ).format(tools=sql.Identifier(self.schema, "aspm_tools"))
            ).fetchone()[0]
            connection.execute(
                sql.SQL(
                    "INSERT INTO {imports} (import_id, owner_user_id, tool_id, source_format, status, started_at) "
                    "VALUES ('stale-import', 'user-a', %s, 'semgrep-api', 'processing', now() - interval '2 days')"
                ).format(imports=sql.Identifier(self.schema, "aspm_imports")),
                (tool_id,),
            )
            connection.execute(
                sql.SQL(
                    "INSERT INTO {syncs} (sync_id, owner_user_id, owner_user_login, connector_key, connector_name, status, started_at) "
                    "VALUES ('stale-sync', 'user-a', 'alice', 'semgrep', 'Semgrep', 'running', now() - interval '2 days')"
                ).format(syncs=sql.Identifier(self.schema, "aspm_connector_syncs"))
            )

        posture = self.repository.posture("user-a")

        tool = next(item for item in posture["tools"] if item["tool_key"] == "semgrep")
        self.assertEqual(tool["last_import_status"], "failed")
        self.assertIn("abandoned", tool["last_import_error"].lower())
        with psycopg.connect(POSTGRES_TEST_DSN, row_factory=psycopg.rows.dict_row) as connection:
            import_status = connection.execute(
                sql.SQL("SELECT status FROM {imports} WHERE import_id = 'stale-import'").format(
                    imports=sql.Identifier(self.schema, "aspm_imports")
                )
            ).fetchone()["status"]
            sync_status = connection.execute(
                sql.SQL("SELECT status FROM {syncs} WHERE sync_id = 'stale-sync'").format(
                    syncs=sql.Identifier(self.schema, "aspm_connector_syncs")
                )
            ).fetchone()["status"]
        self.assertEqual(import_status, "failed")
        self.assertEqual(sync_status, "failed")

    def test_remediation_policy_recalculates_policy_dates_but_keeps_manual_due_dates(self) -> None:
        document = parse_finding_document(
            {
                "format": "generic",
                "tool": {"key": "timeline-test", "name": "Timeline Test"},
                "findings": [
                    {"id": "timeline-1", "title": "Timeline finding", "severity": "critical"}
                ],
            }
        )
        self.repository.ingest("user-a", "alice", document)
        finding = self.repository.search_findings("user-a")["rows"][0]
        initial_due = finding["due_at"]

        result = self.repository.update_remediation_policy(
            "user-a", {"critical": 2, "high": 30, "medium": 90, "low": 180, "info": 365}
        )
        policy_due = self.repository.finding_detail("user-a", finding["finding_id"])["finding"]["due_at"]
        manual_due = "2030-01-01T00:00:00Z"
        self.repository.update_finding(
            "user-a", "alice", finding["finding_id"], "triaged", due_at=manual_due
        )
        self.repository.update_remediation_policy(
            "user-a", {"critical": 1, "high": 30, "medium": 90, "low": 180, "info": 365}
        )
        preserved_due = self.repository.finding_detail("user-a", finding["finding_id"])["finding"]["due_at"]

        self.assertEqual(result["updatedFindings"], 1)
        self.assertNotEqual(policy_due, initial_due)
        self.assertEqual(preserved_due, "2030-01-01T00:00:00+00:00")

    def test_mobile_identifier_and_web_domain_correlation_build_asset_risk(self) -> None:
        with psycopg.connect(POSTGRES_TEST_DSN) as connection:
            connection.execute(
                sql.SQL(
                    "UPDATE {branches} SET mobile_identifier = 'com.example.payments' WHERE branch_inventory_id = %s"
                ).format(branches=sql.Identifier(self.schema, "branch_inventory")),
                (self.branch_inventory_id,),
            )
        mobile_document = parse_finding_document(
            {
                "format": "generic",
                "tool": {
                    "key": "nowsecure",
                    "name": "NowSecure",
                    "type": "mobile_security",
                },
                "findings": [
                    {
                        "id": "mobile-1",
                        "title": "Hardcoded payment service credential",
                        "severity": "high",
                        "cwe": "CWE-798",
                        "application_identifier": "com.example.payments",
                        "dataInteractions": [
                            {
                                "dataType": "payment_card_data",
                                "confidence": 0.98,
                                "source": "scanner",
                                "evidence": "PCI data",
                            }
                        ],
                    }
                ],
            }
        )
        web_document = parse_finding_document(
            {
                "format": "generic",
                "tool": {"key": "invicti", "name": "Invicti", "type": "dast"},
                "findings": [
                    {
                        "id": "web-1",
                        "title": "Personal data exposure",
                        "severity": "critical",
                        "web_url": "https://payments.example.test/account",
                        "cwe": "CWE-200",
                    }
                ],
            }
        )

        mobile_result = self.repository.ingest("user-a", "alice", mobile_document)
        web_result = self.repository.ingest("user-a", "alice", web_document)

        self.assertEqual(mobile_result["linkedFindings"], 1)
        self.assertEqual(web_result["linkedFindings"], 1)
        profile = self.repository.asset_profile("user-a", self.branch_inventory_id)
        self.assertIn("payment_card_data", profile["data_types"])
        self.assertIn("credentials", profile["data_types"])
        self.assertGreater(profile["risk_score"], 0)
        assets = self.repository.asset_risks(
            "user-a", data_types=["payment_card_data"]
        )
        self.assertEqual(assets["total"], 1)
        self.assertEqual(
            assets["rows"][0]["branch_inventory_id"], self.branch_inventory_id
        )

    def test_asset_without_findings_receives_an_explainable_baseline_risk(self) -> None:
        assets = self.repository.asset_risks("user-a")

        self.assertEqual(assets["total"], 1)
        self.assertGreater(assets["rows"][0]["context_score"], 0)
        self.assertGreater(assets["rows"][0]["risk_score"], 0)
        profile = self.repository.asset_profile("user-a", self.branch_inventory_id)
        self.assertEqual(profile["active_findings"], 0)
        self.assertEqual(profile["risk_factors"][2]["factor"], "asset_context")

    def test_changed_correlation_refreshes_previous_and_current_assets(self) -> None:
        with psycopg.connect(POSTGRES_TEST_DSN) as connection:
            repository_id = connection.execute(
                sql.SQL(
                    "INSERT INTO {repositories} (owner_user_id, owner_user_login, provider, organization, project, repo_name, web_url) VALUES ('user-a', 'alice', 'github-enterprise', 'ExampleEngineering', '', 'payments-worker', 'https://github.com/ExampleEngineering/payments-worker') RETURNING repository_id"
                ).format(repositories=sql.Identifier(self.schema, "repositories"))
            ).fetchone()[0]
            current_asset_id = connection.execute(
                sql.SQL(
                    """
                    INSERT INTO {branches} (
                        repository_id, scan_id, owner_user_id, owner_user_login,
                        branch_name, inventory_name, primary_language, last_updated,
                        confidence, score, row_data, detection_evidence, scan_started_at
                    ) VALUES (%s, 'scan-1', 'user-a', 'alice', 'main',
                              'Payments Worker', 'Python', now(), 'high', 90,
                              '{{}}'::jsonb, '{{}}'::jsonb, now())
                    RETURNING branch_inventory_id
                    """
                ).format(branches=sql.Identifier(self.schema, "branch_inventory")),
                (repository_id,),
            ).fetchone()[0]
        initial = FindingDocument(
            tool_key="semgrep",
            tool_name="Semgrep",
            tool_type="sast",
            source_format="semgrep-api",
            findings=(
                FindingInput(
                    external_id="stable-finding",
                    fingerprint_hint="stable-fingerprint",
                    title="SQL injection",
                    severity="high",
                    location=SourceLocation(
                        provider="github-enterprise",
                        organization="ExampleEngineering",
                        repository="payments-api",
                        branch="main",
                    ),
                ),
            ),
        )
        corrected = FindingDocument(
            tool_key="semgrep",
            tool_name="Semgrep",
            tool_type="sast",
            source_format="semgrep-api",
            findings=(
                FindingInput(
                    external_id="stable-finding",
                    fingerprint_hint="stable-fingerprint",
                    title="SQL injection",
                    severity="high",
                    location=SourceLocation(
                        provider="github-enterprise",
                        organization="ExampleEngineering",
                        repository="payments-worker",
                        branch="main",
                    ),
                ),
            ),
        )

        self.repository.ingest("user-a", "alice", initial)
        self.repository.ingest("user-a", "alice", corrected)

        previous = self.repository.asset_profile(
            "user-a", self.branch_inventory_id
        )
        current = self.repository.asset_profile("user-a", current_asset_id)
        self.assertEqual(previous["active_findings"], 0)
        self.assertEqual(current["active_findings"], 1)

    def test_streamed_batches_share_one_import_and_reconcile_after_completion(
        self,
    ) -> None:
        location = SourceLocation(
            provider="github-enterprise",
            organization="ExampleEngineering",
            repository="payments-api",
            branch="main",
        )
        previous = FindingDocument(
            tool_key="semgrep",
            tool_name="Semgrep",
            tool_type="sast",
            source_format="semgrep-api",
            findings=(
                FindingInput(
                    external_id="previous",
                    title="Previous finding",
                    severity="medium",
                    location=location,
                ),
            ),
        )
        self.repository.ingest("user-a", "alice", previous)
        first_batch = FindingDocument(
            tool_key="semgrep",
            tool_name="Semgrep",
            tool_type="sast",
            source_format="semgrep-api",
            findings=(
                FindingInput(
                    external_id="current",
                    title="Hard-coded API key",
                    severity="high",
                    location=location,
                    cwes=("CWE-798",),
                    cves=("CVE-2026-1234",),
                ),
            ),
            scanned_targets=(location,),
            metadata={"page": 0},
        )
        final_batch = FindingDocument(
            tool_key="semgrep",
            tool_name="Semgrep",
            tool_type="sast",
            source_format="semgrep-api",
            findings=(),
            complete_snapshot=True,
            metadata={"recordsRead": 1},
        )
        active_during_import: list[int] = []

        def batches():
            yield first_batch
            active_during_import.append(
                self.repository.asset_profile(
                    "user-a", self.branch_inventory_id
                )["active_findings"]
            )
            yield final_batch

        result = self.repository.ingest_batches("user-a", "alice", batches())

        self.assertEqual(result["findings"], 1)
        self.assertEqual(result["inserted"], 1)
        self.assertEqual(result["resolved"], 1)
        self.assertEqual(result["metadata"]["recordsRead"], 1)
        self.assertEqual(active_during_import, [2])
        self.assertEqual(
            self.repository.asset_profile(
                "user-a", self.branch_inventory_id
            )["active_findings"],
            1,
        )
        resolved = self.repository.search_findings(
            "user-a", filters={"statuses": ["resolved"]}
        )
        self.assertEqual(resolved["total"], 1)
        with psycopg.connect(
            POSTGRES_TEST_DSN, row_factory=psycopg.rows.dict_row
        ) as connection:
            imported = connection.execute(
                sql.SQL(
                    "SELECT status, finding_count, complete_snapshot FROM {imports} WHERE import_id = %s"
                ).format(imports=sql.Identifier(self.schema, "aspm_imports")),
                (result["importId"],),
            ).fetchone()
            relationships = connection.execute(
                sql.SQL(
                    """
                    SELECT
                        (SELECT count(*) FROM {identifiers}) AS identifiers,
                        (SELECT count(*) FROM {data_types}) AS data_types,
                        (SELECT count(*) FROM {import_findings}
                         WHERE import_id = %s) AS import_findings,
                        (SELECT count(*) FROM {events}
                         WHERE event_type = 'created') AS created_events
                    """
                ).format(
                    identifiers=sql.Identifier(
                        self.schema, "aspm_finding_identifiers"
                    ),
                    data_types=sql.Identifier(
                        self.schema, "aspm_finding_data_types"
                    ),
                    import_findings=sql.Identifier(
                        self.schema, "aspm_import_findings"
                    ),
                    events=sql.Identifier(self.schema, "aspm_finding_events"),
                ),
                (result["importId"],),
            ).fetchone()
        self.assertEqual(imported["status"], "completed")
        self.assertEqual(imported["finding_count"], 1)
        self.assertTrue(imported["complete_snapshot"])
        self.assertEqual(relationships["identifiers"], 2)
        self.assertEqual(relationships["data_types"], 2)
        self.assertEqual(relationships["import_findings"], 1)
        self.assertEqual(relationships["created_events"], 2)


if __name__ == "__main__":
    unittest.main()
