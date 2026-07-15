import json
import stat
import tempfile
import time
import unittest
from datetime import datetime, timedelta, timezone
from pathlib import Path
from unittest.mock import Mock, patch

import appsec_scan_router as scanner
import appsec_scan_router.azure as azure_module
import appsec_scan_router.github as github_module
import appsec_scan_router.scanner as scanner_module
import appsec_scan_router.source_discovery as source_discovery_module
import appsec_scan_router.ui as ui_module
import application_inventory_service
import ado_mobile_scanner
import mobile_scanner
from appsec_scan_router.auth import (
    AuthManager,
    AuthenticatedUser,
    CredentialStore,
    GitHubEnterpriseOAuthConfig,
    GoogleOAuthConfig,
    TestLoginConfig,
    expired_session_cookie,
    session_cookie,
)
from appsec_scan_router.github import (
    GitHubAppCredentials,
    GitHubAppTokenProvider,
    normalize_github_api_url,
)
from appsec_scan_router.postgres import (
    POSTGRES_COLUMNS,
    PRIMARY_KEY_COLUMNS,
    normalize_search_query,
    sanitize_observability_message,
    search_tokens,
)
from appsec_scan_router.ui import default_ui_config
from openpyxl import load_workbook


def workbook_value(sheet, field_name, row):
    headers = [cell.value for cell in sheet[1]]
    return sheet.cell(row=row, column=headers.index(field_name) + 1).value


class PublicApiTests(unittest.TestCase):
    def test_package_api_is_importable(self):
        self.assertIs(scanner.ScanConfig, ado_mobile_scanner.ScanConfig)
        self.assertIs(scanner.ScanConfig, mobile_scanner.ScanConfig)
        self.assertIs(scanner.ScanConfig, application_inventory_service.ScanConfig)
        self.assertTrue(callable(scanner.scan))
        self.assertTrue(callable(scanner.scan_to_reports))
        self.assertTrue(callable(scanner.detect_mobile_repo))
        self.assertTrue(callable(scanner.detect_inventory_repo))
        self.assertTrue(callable(scanner.discover_web_domains))
        self.assertTrue(callable(scanner.ApplicationInventoryService))
        self.assertIs(
            scanner.ApplicationInventoryService,
            application_inventory_service.ApplicationInventoryService,
        )
        self.assertTrue(callable(scanner.AppSecInventoryService))
        self.assertTrue(callable(scanner.AppSecScanRouter))
        self.assertTrue(callable(scanner.GitHubEnterpriseClient))
        self.assertTrue(callable(scanner.PostgresInventoryWriter))
        self.assertTrue(callable(scanner.database_status))
        self.assertTrue(callable(scanner.export_inventory_csv))
        self.assertTrue(callable(scanner.export_inventory_json))
        self.assertTrue(callable(scanner.export_inventory_rows))
        self.assertTrue(callable(scanner.export_inventory_xlsx))
        self.assertTrue(callable(scanner.search_inventory))
        self.assertIn("mobile_app", scanner.KNOWN_INVENTORY_TYPES)
        self.assertEqual(scanner.APPLICATION_TYPE_LABELS["ai_enabled"], "AI-enabled")
        self.assertIn("ml_enabled", scanner.KNOWN_INVENTORY_TYPES)
        self.assertEqual(scanner.APPLICATION_TYPE_LABELS["ml_enabled"], "ML-enabled")
        self.assertEqual(scanner.DEFAULT_POSTGRES_SCHEMA, "application_inventory")
        self.assertEqual(scanner.DEFAULT_GITHUB_APP_ID, "")
        self.assertEqual(scanner.DEFAULT_GITHUB_APP_INSTALLATION_ID, "")
        self.assertTrue(callable(scanner.parse_ado_org_pat_values))
        self.assertTrue(callable(scanner.ado_org_pats_to_json))
        self.assertEqual(
            scanner.parse_github_urls(
                ["https://github.com/global-snt", "security-team"]
            ),
            ("global-snt", "security-team"),
        )
        self.assertTrue(callable(scanner.parse_source_target_filter_values))
        self.assertTrue(callable(scanner.source_target_filters_to_json))
        self.assertIs(scanner.SourceTargetFilter, ado_mobile_scanner.SourceTargetFilter)
        org_pats = scanner.parse_ado_org_pat_values("FabrikamCloud=pat-a")
        self.assertEqual(
            [(item.org, item.pat) for item in org_pats], [("FabrikamCloud", "pat-a")]
        )
        target_filters = scanner.parse_source_target_filter_values(
            "FabrikamCloud=Go_To_Market"
        )
        self.assertEqual(
            [(item.org, item.project) for item in target_filters],
            [("FabrikamCloud", "Go_To_Market")],
        )

    def test_report_file_stem_labels_selected_application_types(self):
        self.assertEqual(
            scanner.report_file_stem("Inventory Scan", ("mobile_app", "api_service")),
            "inventory_scan_mobile_app_api_service",
        )
        self.assertEqual(
            scanner.report_file_stem("Inventory Scan", ()), "inventory_scan_all_types"
        )

    def test_azure_throttle_honors_retry_after(self):
        class Response:
            headers = {"Retry-After": "1"}
            status_code = 200

        throttle = azure_module.AzureDevOpsThrottle(
            requests_per_second=0, low_remaining_backoff_seconds=2
        )
        before = time.monotonic()
        throttle.observe(Response())
        self.assertGreaterEqual(throttle.block_until, before + 0.9)

    def test_azure_throttle_backs_off_when_rate_limit_remaining_is_empty(self):
        class Response:
            headers = {"X-RateLimit-Remaining": "0"}
            status_code = 200

        throttle = azure_module.AzureDevOpsThrottle(
            requests_per_second=0, low_remaining_backoff_seconds=2
        )
        before = time.monotonic()
        throttle.observe(Response())
        self.assertGreaterEqual(throttle.block_until, before + 1.9)

    def test_github_throttle_reserves_the_remaining_request_budget(self):
        class Response:
            headers = {
                "X-RateLimit-Remaining": "25",
                "X-RateLimit-Reset": str(time.time() + 60),
            }
            status_code = 200

        throttle = github_module.GitHubThrottle(
            requests_per_second=0, rate_limit_reserve=50
        )
        before = time.monotonic()
        throttle.observe(Response())
        self.assertGreaterEqual(throttle.block_until, before + 59)

    def test_content_selection_excludes_dependency_artifacts(self):
        self.assertTrue(scanner.should_fetch_content("/services/api/package.json"))
        self.assertFalse(
            scanner.should_fetch_content("/services/api/node_modules/pkg/package.json")
        )
        self.assertFalse(
            scanner.should_fetch_content("/services/api/package-lock.json")
        )


class AuthTests(unittest.TestCase):
    def test_credential_store_encrypts_saved_tokens(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            store = CredentialStore(Path(tmpdir))

            store.save_token("user-1", "azure-devops", "secret-token")

            self.assertEqual(store.token("user-1", "azure-devops"), "secret-token")
            self.assertTrue(store.statuses("user-1")["azure-devops"])
            encrypted = (Path(tmpdir) / "credentials.json.enc").read_bytes()
            self.assertNotIn(b"secret-token", encrypted)

    def test_google_oauth_config_reads_environment(self):
        with patch.dict(
            "os.environ",
            {
                "APPLICATION_INVENTORY_SERVICE_GOOGLE_CLIENT_ID": "google-client",
                "APPLICATION_INVENTORY_SERVICE_GOOGLE_CLIENT_SECRET": "google-secret",
            },
            clear=False,
        ):
            config = GoogleOAuthConfig.from_env()

        self.assertTrue(config.enabled)
        self.assertEqual(config.client_id, "google-client")
        self.assertEqual(config.scope, "openid email profile")

    def test_github_enterprise_oauth_config_derives_endpoints(self):
        with patch.dict(
            "os.environ",
            {
                "APPLICATION_INVENTORY_SERVICE_GHE_BASE_URL": "https://ghe.example.com/api/v3",
                "APPLICATION_INVENTORY_SERVICE_GHE_CLIENT_ID": "ghe-client",
                "APPLICATION_INVENTORY_SERVICE_GHE_CLIENT_SECRET": "ghe-secret",
            },
            clear=True,
        ):
            config = GitHubEnterpriseOAuthConfig.from_env()

        self.assertTrue(config.enabled)
        self.assertEqual(
            config.authorize_url, "https://ghe.example.com/login/oauth/authorize"
        )
        self.assertEqual(
            config.token_url, "https://ghe.example.com/login/oauth/access_token"
        )
        self.assertEqual(config.user_url, "https://ghe.example.com/api/v3/user")
        self.assertEqual(config.scope, "read:user read:org")

    def test_github_enterprise_oauth_config_rejects_insecure_base_url_by_default(self):
        with patch.dict(
            "os.environ",
            {
                "APPLICATION_INVENTORY_SERVICE_GHE_BASE_URL": "http://ghe.example.com",
                "APPLICATION_INVENTORY_SERVICE_GHE_CLIENT_ID": "ghe-client",
                "APPLICATION_INVENTORY_SERVICE_GHE_CLIENT_SECRET": "ghe-secret",
            },
            clear=True,
        ):
            config = GitHubEnterpriseOAuthConfig.from_env()

        self.assertFalse(config.enabled)

    def test_test_login_config_reads_environment(self):
        with patch.dict(
            "os.environ",
            {
                "APPLICATION_INVENTORY_SERVICE_TEST_LOGIN_ENABLED": "true",
                "APPLICATION_INVENTORY_SERVICE_TEST_USER_ID": "local-user",
                "APPLICATION_INVENTORY_SERVICE_TEST_USER_LOGIN": "local.user@example.test",
                "APPLICATION_INVENTORY_SERVICE_TEST_USER_NAME": "Local User",
            },
            clear=False,
        ):
            config = TestLoginConfig.from_env()

        user = config.user()

        self.assertTrue(config.enabled)
        self.assertEqual(user.id, "local-user")
        self.assertEqual(user.login, "local.user@example.test")
        self.assertEqual(user.name, "Local User")
        self.assertEqual(user.provider, "test")

    def test_test_login_config_defaults_to_local_user_enabled(self):
        with patch.dict("os.environ", {}, clear=True):
            config = TestLoginConfig.from_env()

        user = config.user()

        self.assertTrue(config.enabled)
        self.assertEqual(user.id, "test-user")
        self.assertEqual(user.login, "test.user@local")
        self.assertEqual(user.provider, "test")

    def test_test_login_config_can_be_disabled(self):
        with patch.dict(
            "os.environ",
            {"APPLICATION_INVENTORY_SERVICE_TEST_LOGIN_ENABLED": "false"},
            clear=True,
        ):
            config = TestLoginConfig.from_env()

        self.assertFalse(config.enabled)

    def test_auth_status_lists_enterprise_and_google_sso(self):
        with (
            tempfile.TemporaryDirectory() as tmpdir,
            patch.dict(
                "os.environ",
                {
                    "APPLICATION_INVENTORY_SERVICE_GHE_BASE_URL": "https://github.enterprise.example",
                    "APPLICATION_INVENTORY_SERVICE_GHE_CLIENT_ID": "github-client",
                    "APPLICATION_INVENTORY_SERVICE_GHE_CLIENT_SECRET": "github-secret",
                    "APPLICATION_INVENTORY_SERVICE_GOOGLE_CLIENT_ID": "google-client",
                    "APPLICATION_INVENTORY_SERVICE_GOOGLE_CLIENT_SECRET": "google-secret",
                    "APPLICATION_INVENTORY_SERVICE_TEST_LOGIN_ENABLED": "true",
                },
                clear=False,
            ),
        ):
            manager = AuthManager(Path(tmpdir))
            status = manager.status(None)

        providers = {provider["id"]: provider for provider in status["authProviders"]}
        self.assertTrue(status["githubEnterpriseLoginEnabled"])
        self.assertTrue(status["googleLoginEnabled"])
        self.assertTrue(status["testLoginEnabled"])
        self.assertEqual(
            providers["github-enterprise"]["startUrl"],
            "/api/auth/github-enterprise/start",
        )
        self.assertEqual(providers["google"]["startUrl"], "/api/auth/google/start")
        self.assertEqual(providers["test"]["startUrl"], "/api/auth/test/start")

    def test_default_ui_config_lists_sso_options(self):
        with (
            tempfile.TemporaryDirectory() as tmpdir,
            patch.dict("os.environ", {}, clear=True),
        ):
            config = default_ui_config(Path(tmpdir))

        provider_ids = [provider["id"] for provider in config["auth"]["authProviders"]]
        application_type_ids = [
            choice["value"] for choice in config["defaults"]["applicationTypeChoices"]
        ]
        self.assertEqual(provider_ids, ["github-enterprise", "google", "test"])
        self.assertIn("googleLoginEnabled", config["auth"])
        self.assertTrue(config["auth"]["testLoginEnabled"])
        self.assertIn("ml_enabled", application_type_ids)
        self.assertEqual(config["defaults"]["activityMode"], "contributors")
        self.assertEqual(config["defaults"]["githubUrls"], [])
        self.assertNotIn("baseUrl", config["defaults"])
        self.assertEqual(config["defaults"]["postgresHost"], "localhost")
        self.assertEqual(config["defaults"]["postgresPassword"], "postgres")
        self.assertEqual(config["defaults"]["postgresSchema"], "application_inventory")

    def test_default_ui_config_reads_backend_github_scope(self):
        with (
            tempfile.TemporaryDirectory() as tmpdir,
            patch.dict(
                "os.environ",
                {
                    "APPLICATION_INVENTORY_GITHUB_URLS": "owner-a,owner-b",
                    "APPLICATION_INVENTORY_GITHUB_REPOSITORIES": "owner-a=payments-api,owner-b=identity-service",
                    "APPLICATION_INVENTORY_GITHUB_API_URL": "https://github.example.com/api/v3",
                },
                clear=True,
            ),
        ):
            config = default_ui_config(Path(tmpdir))

        self.assertEqual(config["defaults"]["githubUrls"], ["owner-a", "owner-b"])
        self.assertEqual(
            config["defaults"]["githubRepositories"],
            ["owner-a=payments-api", "owner-b=identity-service"],
        )
        self.assertNotIn("baseUrl", config["defaults"])

    def test_github_enterprise_session_stores_token_for_scan(self):
        with (
            tempfile.TemporaryDirectory() as tmpdir,
            patch.dict("os.environ", {}, clear=True),
        ):
            manager = AuthManager(Path(tmpdir))
            user = AuthenticatedUser(
                id="ghe-user", login="ghe.user", provider="github-enterprise"
            )
            record = manager.create_session(user, provider_token="ghe-access-token")

            payload = manager.apply_credentials(
                {"provider": "github-enterprise"}, record
            )

        self.assertEqual(payload["token"], "ghe-access-token")

    def test_session_cookies_include_security_attributes(self):
        cookie = session_cookie("session-id", secure=True)
        expired = expired_session_cookie(secure=True)

        self.assertIn("HttpOnly", cookie)
        self.assertIn("SameSite=Lax", cookie)
        self.assertIn("Secure", cookie)
        self.assertIn("HttpOnly", expired)
        self.assertIn("SameSite=Lax", expired)
        self.assertIn("Secure", expired)

    def test_sessions_survive_restart_in_encrypted_state(self):
        with (
            tempfile.TemporaryDirectory() as tmpdir,
            patch.dict("os.environ", {}, clear=True),
        ):
            reports_root = Path(tmpdir)
            first_manager = AuthManager(reports_root)
            created = first_manager.create_test_session()
            session_path = (
                reports_root
                / ".application_inventory_service"
                / "sessions.json.enc"
            )

            second_manager = AuthManager(reports_root)
            recovered = second_manager.sessions.get(created.id)

            self.assertIsNotNone(recovered)
            self.assertEqual(recovered.user, created.user)
            self.assertEqual(recovered.csrf_token, created.csrf_token)
            encrypted = session_path.read_bytes()
            self.assertNotIn(created.id.encode(), encrypted)
            self.assertNotIn(created.csrf_token.encode(), encrypted)
            self.assertEqual(stat.S_IMODE(session_path.stat().st_mode), 0o600)

            second_manager.logout(created.id)
            third_manager = AuthManager(reports_root)
            self.assertIsNone(third_manager.sessions.get(created.id))

    def test_security_headers_include_browser_defenses(self):
        with patch.dict(
            "os.environ",
            {"APPLICATION_INVENTORY_SERVICE_COOKIE_SECURE": "true"},
            clear=False,
        ):
            headers = ui_module.security_headers()

        self.assertIn("Content-Security-Policy", headers)
        self.assertIn("frame-ancestors 'none'", headers["Content-Security-Policy"])
        self.assertEqual(headers["X-Content-Type-Options"], "nosniff")
        self.assertEqual(headers["X-Frame-Options"], "DENY")
        self.assertIn("Strict-Transport-Security", headers)

    def test_public_url_rejects_credentials_and_insecure_https_mode(self):
        with patch.dict(
            "os.environ",
            {"APPLICATION_INVENTORY_SERVICE_COOKIE_SECURE": "true"},
            clear=False,
        ):
            with self.assertRaises(ValueError):
                ui_module.safe_public_url("http://inventory.example.com")
            with self.assertRaises(ValueError):
                ui_module.safe_public_url("https://user:pass@inventory.example.com")

        self.assertEqual(
            ui_module.safe_public_url("https://inventory.example.com/app"),
            "https://inventory.example.com",
        )

    def test_request_base_url_rejects_header_injection_hosts(self):
        self.assertEqual(
            ui_module.safe_request_base_url("https", "inventory.example.com"),
            "https://inventory.example.com",
        )
        with self.assertRaises(ValueError):
            ui_module.safe_request_base_url(
                "https", "inventory.example.com\r\nX-Test: true"
            )
        with self.assertRaises(ValueError):
            ui_module.safe_request_base_url("https", "inventory.example.com/path")

    def test_github_api_url_normalization_rejects_unsafe_urls(self):
        self.assertEqual(
            normalize_github_api_url("github.example.com"),
            "https://github.example.com/api/v3",
        )
        with self.assertRaises(ValueError):
            normalize_github_api_url("https://token@github.example.com/api/v3")
        with self.assertRaises(ValueError):
            normalize_github_api_url("http://github.example.com/api/v3")

    def test_github_api_url_respects_allowed_host_list(self):
        with patch.dict(
            "os.environ",
            {
                "APPLICATION_INVENTORY_SERVICE_ALLOWED_GITHUB_HOSTS": "github.allowed.example"
            },
            clear=False,
        ):
            self.assertEqual(
                normalize_github_api_url("https://github.allowed.example"),
                "https://github.allowed.example/api/v3",
            )
            with self.assertRaises(ValueError):
                normalize_github_api_url("https://github.blocked.example")


class MultiOrganizationScanTests(unittest.TestCase):
    def test_scan_dispatches_each_github_url(self):
        config = scanner.ScanConfig(
            org="global-snt",
            pat="github-token",
            project=None,
            out_dir=Path("reports"),
            out_prefix="scan",
            max_workers=1,
            content_workers=1,
            max_commits_per_repo=0,
            timeout_seconds=30,
            min_confidence="low",
            provider="github-enterprise",
            base_url="https://api.github.com",
            github_urls=("global-snt", "security-team"),
        )

        def fake_scan(source_config, on_result=None):
            return [
                {"provider": source_config.provider, "organization": source_config.org}
            ]

        with patch.object(
            scanner_module, "scan_single_org", side_effect=fake_scan
        ) as scan_single_org:
            results = scanner.scan(config)

        self.assertEqual(
            [call.args[0].org for call in scan_single_org.call_args_list],
            ["global-snt", "security-team"],
        )
        self.assertEqual(
            [row["organization"] for row in results], ["global-snt", "security-team"]
        )

    def test_scan_mixed_dispatches_all_configured_sources(self):
        config = scanner.ScanConfig(
            org="FabrikamGH",
            pat="github-token",
            project=None,
            out_dir=Path("reports"),
            out_prefix="scan",
            max_workers=1,
            content_workers=1,
            max_commits_per_repo=0,
            timeout_seconds=30,
            min_confidence="low",
            provider="mixed",
            base_url="https://github.fabrikam.example/api/v3",
            ado_org_pats=(scanner.AzureDevOpsOrgPat("FabrikamADO", "ado-token"),),
        )

        def fake_scan(source_config, on_result=None):
            return [
                {
                    "provider": source_config.provider,
                    "organization": source_config.org,
                    "project": "",
                    "repo_name": "",
                    "branch_name": "",
                }
            ]

        with patch.object(
            scanner_module, "scan_single_org", side_effect=fake_scan
        ) as scan_single_org:
            results = scanner.scan(config)

        self.assertEqual(
            [
                (call.args[0].provider, call.args[0].org)
                for call in scan_single_org.call_args_list
            ],
            [("azure-devops", "FabrikamADO"), ("github-enterprise", "FabrikamGH")],
        )
        self.assertEqual(
            {row["provider"] for row in results}, {"azure-devops", "github-enterprise"}
        )

    def test_mixed_scan_streams_both_sources_to_one_report_set(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            config = scanner.ScanConfig(
                org="FabrikamGH",
                pat="github-token",
                project=None,
                out_dir=Path(tmpdir),
                out_prefix="scan",
                max_workers=1,
                content_workers=1,
                max_commits_per_repo=0,
                timeout_seconds=30,
                min_confidence="low",
                provider="mixed",
                base_url="https://github.fabrikam.example/api/v3",
                ado_org_pats=(scanner.AzureDevOpsOrgPat("FabrikamADO", "ado-token"),),
            )

            def fake_scan(source_config, on_result=None):
                result = {
                    "provider": source_config.provider,
                    "organization": source_config.org,
                    "project": "Project",
                    "repo_name": "Repo",
                    "branch_name": "main",
                    "branch_age_bucket": scanner.ACTIVE_SHEET_NAME,
                    "semgrep_target": f"{source_config.provider}:target",
                }
                if on_result:
                    on_result(result)
                return [result]

            with patch.object(scanner_module, "scan_single_org", side_effect=fake_scan):
                results, xlsx_path, semgrep_path, sonarqube_path = (
                    scanner.scan_to_reports(config)
                )

            workbook = load_workbook(xlsx_path)
            self.assertEqual(len(results), 2)
            self.assertEqual(workbook[scanner.ACTIVE_SHEET_NAME].max_row, 3)
            self.assertEqual(
                workbook_value(workbook[scanner.ACTIVE_SHEET_NAME], "provider", 2),
                "azure-devops",
            )
            self.assertEqual(
                workbook_value(workbook[scanner.ACTIVE_SHEET_NAME], "provider", 3),
                "github-enterprise",
            )
            self.assertEqual(
                len(semgrep_path.read_text(encoding="utf-8").splitlines()), 2
            )
            self.assertEqual(
                len(sonarqube_path.read_text(encoding="utf-8").splitlines()), 3
            )

    def test_scan_dispatches_each_ado_org_pat(self):
        config = scanner.ScanConfig(
            org="",
            pat="",
            project=None,
            out_dir=Path("reports"),
            out_prefix="scan",
            max_workers=1,
            content_workers=1,
            max_commits_per_repo=0,
            timeout_seconds=30,
            min_confidence="low",
            ado_org_pats=(
                scanner.AzureDevOpsOrgPat("FabrikamCloud", "pat-a"),
                scanner.AzureDevOpsOrgPat("ContosoApps", "pat-b"),
            ),
        )

        def fake_scan(org_config, on_result=None):
            return [
                {
                    "organization": org_config.org,
                    "project": "",
                    "repo_name": "",
                    "branch_name": "",
                }
            ]

        with patch.object(
            scanner_module, "scan_single_org", side_effect=fake_scan
        ) as scan_single_org:
            results = scanner.scan(config)

        self.assertEqual(
            [call.args[0].org for call in scan_single_org.call_args_list],
            ["FabrikamCloud", "ContosoApps"],
        )
        self.assertEqual(
            [call.args[0].pat for call in scan_single_org.call_args_list],
            ["pat-a", "pat-b"],
        )
        self.assertEqual(
            [row["organization"] for row in results], ["ContosoApps", "FabrikamCloud"]
        )

    def test_scan_skips_multi_orgs_without_matching_target_filters(self):
        config = scanner.ScanConfig(
            org="",
            pat="",
            project=None,
            out_dir=Path("reports"),
            out_prefix="scan",
            max_workers=1,
            content_workers=1,
            max_commits_per_repo=0,
            timeout_seconds=30,
            min_confidence="low",
            ado_org_pats=(
                scanner.AzureDevOpsOrgPat("FabrikamCloud", "pat-a"),
                scanner.AzureDevOpsOrgPat("ContosoApps", "pat-b"),
            ),
            target_filters=(scanner.SourceTargetFilter("ContosoApps", "Payments"),),
        )

        def fake_scan(org_config, on_result=None):
            return [
                {
                    "organization": org_config.org,
                    "project": org_config.target_filters[0].project,
                    "repo_name": "",
                    "branch_name": "",
                }
            ]

        with patch.object(
            scanner_module, "scan_single_org", side_effect=fake_scan
        ) as scan_single_org:
            results = scanner.scan(config)

        self.assertEqual(
            [call.args[0].org for call in scan_single_org.call_args_list],
            ["ContosoApps"],
        )
        self.assertEqual(results[0]["project"], "Payments")


class ProviderClientTests(unittest.TestCase):
    def test_normalizes_github_enterprise_api_urls(self):
        self.assertEqual(scanner.normalize_github_api_url(""), "https://api.github.com")
        self.assertEqual(
            scanner.normalize_github_api_url("https://github.fabrikam.example"),
            "https://github.fabrikam.example/api/v3",
        )

    def test_github_tree_scan_uses_the_branch_ref_directly(self):
        client = object.__new__(github_module.GitHubEnterpriseClient)
        client.get_json = Mock(
            return_value={"tree": [{"path": "package.json", "type": "blob"}]}
        )

        items = client.list_repo_items("owner", "owner/repo", "release/prod")

        self.assertEqual(items, [{"path": "/package.json"}])
        path, params = client.get_json.call_args.args
        self.assertEqual(path, "/repos/owner/repo/git/trees/release%2Fprod")
        self.assertEqual(params, {"recursive": "1"})
        self.assertEqual(
            scanner.normalize_github_api_url("https://github.fabrikam.example/api/v3"),
            "https://github.fabrikam.example/api/v3",
        )

    def test_github_app_token_provider_signs_jwt_and_caches_installation_token(self):
        private_key = "-----BEGIN PRIVATE KEY-----\nfake-key\n-----END PRIVATE KEY-----"
        credentials = GitHubAppCredentials("123", "456", private_key)
        provider = GitHubAppTokenProvider(
            "https://github.fabrikam.example/api/v3", credentials, 10
        )
        response = Mock()
        response.json.return_value = {
            "token": "installation-token",
            "expires_at": (datetime.now(timezone.utc) + timedelta(hours=1)).isoformat(),
        }
        provider._session.post = Mock(return_value=response)
        provider._app_jwt = Mock(return_value="app-jwt")

        self.assertEqual(provider.token(), "installation-token")
        self.assertEqual(provider.token(), "installation-token")
        provider._session.post.assert_called_once()
        request_url = provider._session.post.call_args.args[0]
        request_headers = provider._session.post.call_args.kwargs["headers"]
        self.assertEqual(
            request_url,
            "https://github.fabrikam.example/api/v3/app/installations/456/access_tokens",
        )
        self.assertEqual(request_headers["Authorization"], "Bearer app-jwt")
        provider.close()

    def test_github_app_credentials_read_environment(self):
        with patch.dict(
            "os.environ",
            {
                "APPLICATION_INVENTORY_GITHUB_APP_ID": "123",
                "APPLICATION_INVENTORY_GITHUB_APP_INSTALLATION_ID": "456",
                "APPLICATION_INVENTORY_GITHUB_APP_PRIVATE_KEY": "-----BEGIN PRIVATE KEY-----\\nfake\\n-----END PRIVATE KEY-----",
            },
            clear=True,
        ):
            credentials = GitHubAppCredentials.from_env()

        self.assertEqual(credentials.app_id, "123")
        self.assertEqual(credentials.installation_id, "456")
        self.assertIn("\n", credentials.private_key)

    def test_github_app_credentials_use_fixed_ids_with_key_file(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            key_path = Path(tmpdir) / "github-app.pem"
            key_path.write_text(
                "-----BEGIN PRIVATE KEY-----\nfake\n-----END PRIVATE KEY-----",
                encoding="utf-8",
            )
            with patch.dict(
                "os.environ",
                {
                    "APPLICATION_INVENTORY_GITHUB_APP_ID": "123",
                    "APPLICATION_INVENTORY_GITHUB_APP_INSTALLATION_ID": "456",
                    "APPLICATION_INVENTORY_GITHUB_APP_PRIVATE_KEY_FILE": str(key_path),
                },
                clear=True,
            ):
                credentials = GitHubAppCredentials.from_env()

        self.assertIsNotNone(credentials)
        self.assertEqual(credentials.app_id, "123")
        self.assertEqual(credentials.installation_id, "456")

    def test_github_app_jwt_uses_string_issuer(self):
        credentials = GitHubAppCredentials(
            "123", "456", "-----BEGIN PRIVATE KEY-----\nfake\n-----END PRIVATE KEY-----"
        )
        provider = GitHubAppTokenProvider("https://api.github.com", credentials, 10)
        with patch.object(
            github_module.jwt, "encode", return_value="app-jwt"
        ) as encode:
            self.assertEqual(provider._app_jwt(), "app-jwt")

        claims = encode.call_args.args[0]
        self.assertEqual(claims["iss"], "123")
        provider.close()

    def test_observability_message_redacts_sensitive_values(self):
        message = "postgresql://postgres:postgres@localhost:5432/postgres --pat secret Bearer token password=secret"

        sanitized = sanitize_observability_message(message)

        self.assertNotIn("postgres:postgres", sanitized)
        self.assertNotIn("secret", sanitized)
        self.assertNotIn("Bearer token", sanitized)

    def test_normalizes_github_commits_for_activity_extraction(self):
        commit = scanner.github_commit_to_activity_commit(
            {
                "commit": {
                    "author": {
                        "name": "Alice Adams",
                        "email": "alice@example.com",
                        "date": "2026-06-01T12:00:00Z",
                    },
                    "committer": {
                        "name": "Build Service",
                        "email": "build@example.com",
                        "date": "2026-06-02T12:00:00Z",
                    },
                }
            }
        )

        self.assertEqual(commit["author"]["name"], "Alice Adams")
        self.assertEqual(commit["author"]["email"], "alice@example.com")
        self.assertEqual(commit["committer"]["date"], "2026-06-02T12:00:00Z")

    def test_parse_args_supports_github_enterprise(self):
        with patch.dict("os.environ", {"GITHUB_TOKEN": "token"}, clear=False):
            config = scanner.parse_args(
                [
                    "--provider",
                    "github-enterprise",
                    "--base-url",
                    "https://github.fabrikam.example/api/v3",
                    "--org",
                    "FabrikamCloud",
                    "--repo",
                    "mobile-app",
                    "--application-type",
                    "mobile_app",
                    "--application-type",
                    "ai_enabled",
                    "--application-type",
                    "ml_enabled",
                    "--out-dir",
                    "reports",
                ]
            )

        self.assertEqual(config.provider, "github-enterprise")
        self.assertEqual(config.base_url, "https://github.fabrikam.example/api/v3")
        self.assertEqual(config.org, "FabrikamCloud")
        self.assertEqual(config.project, "mobile-app")
        self.assertEqual(config.pat, "token")
        self.assertEqual(
            config.application_types, ("mobile_app", "ai_enabled", "ml_enabled")
        )

    def test_parse_args_supports_github_app_authentication(self):
        with patch.dict(
            "os.environ",
            {
                "APPLICATION_INVENTORY_GITHUB_APP_ID": "123",
                "APPLICATION_INVENTORY_GITHUB_APP_INSTALLATION_ID": "456",
                "APPLICATION_INVENTORY_GITHUB_APP_PRIVATE_KEY": "-----BEGIN PRIVATE KEY-----\\nfake\\n-----END PRIVATE KEY-----",
            },
            clear=True,
        ):
            config = scanner.parse_args(
                [
                    "--provider",
                    "github-enterprise",
                    "--base-url",
                    "https://github.fabrikam.example/api/v3",
                    "--org",
                    "FabrikamCloud",
                ]
            )

        self.assertEqual(config.pat, "")
        self.assertEqual(config.github_app_id, "123")
        self.assertEqual(config.github_app_installation_id, "456")
        self.assertIn("BEGIN PRIVATE KEY", config.github_app_private_key)

    def test_create_source_client_supports_github_enterprise(self):
        config = scanner.ScanConfig(
            org="FabrikamCloud",
            pat="token",
            project="mobile-app",
            out_dir=Path("reports"),
            out_prefix="scan",
            max_workers=1,
            content_workers=1,
            max_commits_per_repo=0,
            timeout_seconds=30,
            min_confidence="low",
            provider="github-enterprise",
            base_url="https://github.fabrikam.example",
        )

        client = scanner.create_source_client(config)
        try:
            self.assertIsInstance(client, scanner.GitHubEnterpriseClient)
            self.assertEqual(client.base_url, "https://github.fabrikam.example/api/v3")
        finally:
            client.close()

    def test_collect_targets_uses_selected_project_filters(self):
        class FakeClient:
            org = "FabrikamCloud"

            def list_projects(self):
                raise AssertionError(
                    "project list should not be loaded when target filters are provided"
                )

            def list_repos(self, project_name):
                return [{"id": f"{project_name}-repo", "name": f"{project_name}-repo"}]

        targets = scanner.collect_targets(
            FakeClient(),
            None,
            (
                scanner.SourceTargetFilter("ContosoApps", "Ignored"),
                scanner.SourceTargetFilter("FabrikamCloud", "Payments"),
                scanner.SourceTargetFilter("", "Shared"),
            ),
        )

        self.assertEqual(
            [(target.project_name, target.repo["name"]) for target in targets],
            [("Payments", "Payments-repo"), ("Shared", "Shared-repo")],
        )

    def test_parse_args_supports_multiple_ado_org_pats(self):
        config = scanner.parse_args(
            [
                "--ado-org-pat",
                "FabrikamCloud=pat-a",
                "--ado-org-pat",
                "ContosoApps=pat-b",
                "--out-dir",
                "reports",
            ]
        )

        self.assertEqual(config.provider, "azure-devops")
        self.assertEqual(config.org, "")
        self.assertEqual(config.pat, "")
        self.assertEqual(
            [(item.org, item.pat) for item in config.ado_org_pats],
            [("FabrikamCloud", "pat-a"), ("ContosoApps", "pat-b")],
        )

    def test_parse_args_supports_repeated_projects_as_target_filters(self):
        with patch.dict("os.environ", {"ADO_PAT": "token"}, clear=False):
            config = scanner.parse_args(
                [
                    "--org",
                    "FabrikamCloud",
                    "--project",
                    "Payments",
                    "--project",
                    "Storefront",
                    "--out-dir",
                    "reports",
                ]
            )

        self.assertIsNone(config.project)
        self.assertEqual(
            [(item.org, item.project) for item in config.target_filters],
            [("", "Payments"), ("", "Storefront")],
        )

    def test_parse_args_supports_explicit_target_filters(self):
        config = scanner.parse_args(
            [
                "--ado-org-pat",
                "FabrikamCloud=pat-a",
                "--ado-org-pat",
                "ContosoApps=pat-b",
                "--target-filter",
                "FabrikamCloud=Payments",
                "--target-filter",
                "ContosoApps=Storefront",
                "--out-dir",
                "reports",
            ]
        )

        self.assertEqual(
            [(item.org, item.project) for item in config.target_filters],
            [("FabrikamCloud", "Payments"), ("ContosoApps", "Storefront")],
        )

    def test_parse_args_supports_ado_org_pats_from_json_environment(self):
        with patch.dict(
            "os.environ",
            {
                "APPSEC_INVENTORY_ADO_ORG_PATS": json.dumps(
                    {"FabrikamCloud": "pat-a", "ContosoApps": "pat-b"}
                )
            },
            clear=False,
        ):
            config = scanner.parse_args(["--out-dir", "reports"])

        self.assertEqual(
            [(item.org, item.pat) for item in config.ado_org_pats],
            [("FabrikamCloud", "pat-a"), ("ContosoApps", "pat-b")],
        )

    def test_parse_args_supports_ado_org_pats_from_json_list_environment(self):
        with patch.dict(
            "os.environ",
            {
                "APPSEC_INVENTORY_ADO_ORG_PATS": json.dumps(
                    [
                        {"org": "FabrikamCloud", "pat": "pat-a"},
                        {"org": "ContosoApps", "pat": "pat-b"},
                    ]
                )
            },
            clear=False,
        ):
            config = scanner.parse_args(["--out-dir", "reports"])

        self.assertEqual(config.org, "")
        self.assertEqual(config.pat, "")
        self.assertEqual(
            [(item.org, item.pat) for item in config.ado_org_pats],
            [("FabrikamCloud", "pat-a"), ("ContosoApps", "pat-b")],
        )


class UiServiceTests(unittest.TestCase):
    def test_ui_hides_server_managed_github_app_credentials(self):
        html = (
            Path(__file__).parents[1]
            / "appsec_scan_router"
            / "ui_static"
            / "index.html"
        ).read_text(encoding="utf-8")
        javascript = (
            Path(__file__).parents[1] / "appsec_scan_router" / "ui_static" / "app.js"
        ).read_text(encoding="utf-8")

        self.assertIn('class="provider-github github-app-status hidden full"', html)
        self.assertIn("Managed by service", html)
        self.assertNotIn("githubAppPrivateKeyFile", html)
        self.assertNotIn("githubAppInstallationId", html)
        self.assertNotIn('value="4255413"', html)
        self.assertNotIn('name="githubAppPrivateKey"', html)
        self.assertNotIn('name="project"', html)
        self.assertNotIn('name="repo"', html)
        self.assertIn('name="githubUrl"', html)
        self.assertNotIn("GitHub API URL", html)
        self.assertNotIn("progressValue", html)
        self.assertNotIn("etaValue", html)
        self.assertIn("log-error", javascript)
        self.assertIn("log-success", javascript)
        self.assertNotIn('name="token"', html)
        self.assertNotIn('name="saveToken"', html)
        self.assertIn('id="adoOrgRequirementBadge">Required</span>', html)
        self.assertNotIn("handleGithubAppPrivateKeyFileChange", javascript)
        self.assertNotIn("DEFAULT_GITHUB_APP_ID", javascript)
        self.assertNotIn(
            '"githubAppPrivateKey"',
            javascript.split("const persistedFields", 1)[1].split("];", 1)[0],
        )
        self.assertNotIn("syncCredentialFields", javascript)
        self.assertNotIn('id="loginGitHubSso"', html)
        self.assertNotIn("/api/auth/github/start", html)
        self.assertNotIn("loginGitHubSso", javascript)
        self.assertIn('id="loginGitHubEnterpriseSso"', html)
        self.assertIn("Local AI inventory assistant", html)
        self.assertIn('id="askInventory" type="button">Ask AI</button>', html)
        self.assertIn('data-database-sort="confidence"', html)
        self.assertIn('id="filterConfidence"', html)
        self.assertIn('id="filterType"', html)
        self.assertIn('id="clearFilterTypes"', html)
        self.assertEqual(html.count('name="databaseFilterType"'), 10)
        self.assertNotIn('<select id="filterType"', html)
        self.assertIn('data-inventory-filter="has-domain"', html)
        for inventory_filter in (
            "mobile-app",
            "web-app",
            "api-service",
            "microservice",
            "middleware",
            "serverless",
            "library",
            "infrastructure",
            "ai-enabled",
            "ml-enabled",
        ):
            self.assertIn(f'data-inventory-filter="{inventory_filter}"', html)
        self.assertIn(
            '".png": "image/png"',
            (Path(__file__).parents[1] / "appsec_scan_router" / "ui.py").read_text(
                encoding="utf-8"
            ),
        )
        self.assertIn(
            '".jpg": "image/jpeg"',
            (Path(__file__).parents[1] / "appsec_scan_router" / "ui.py").read_text(
                encoding="utf-8"
            ),
        )
        self.assertIn(
            '".svg": "image/svg+xml"',
            (Path(__file__).parents[1] / "appsec_scan_router" / "ui.py").read_text(
                encoding="utf-8"
            ),
        )

    def test_normalize_scan_config_requires_org(self):
        with self.assertRaises(ValueError):
            scanner.normalize_scan_config({"provider": "azure-devops"})
        with self.assertRaises(ValueError):
            scanner.normalize_scan_config(
                {"provider": "azure-devops", "org": "FabrikamCloud"}
            )

    def test_normalize_scan_config_defaults_github_url_and_api_url(self):
        with patch.dict(
            "os.environ",
            {"APPLICATION_INVENTORY_GITHUB_URLS": "global-snt"},
            clear=False,
        ):
            config = scanner.normalize_scan_config({"provider": "github-enterprise"})

        self.assertEqual(config["githubUrls"], ["global-snt"])
        self.assertEqual(config["org"], "global-snt")
        self.assertEqual(config["baseUrl"], "https://api.github.com")

    def test_build_scan_command_for_github_enterprise(self):
        with patch.dict(
            "os.environ",
            {
                "APPLICATION_INVENTORY_GITHUB_API_URL": "https://github.fabrikam.example/api/v3"
            },
            clear=False,
        ):
            config = scanner.normalize_scan_config(
                {
                    "provider": "github-enterprise",
                    "org": "FabrikamCloud",
                    "repo": "mobile-app",
                    "outPrefix": "inventory scan",
                    "applicationTypes": ["mobile_app", "ai_enabled", "ml_enabled"],
                    "minConfidence": "medium",
                    "activityMode": "latest",
                    "storeLookup": True,
                }
            )

        command = scanner.build_scan_command(config, Path("/reports/scan-1"))

        self.assertIn("--provider", command)
        self.assertIn("github-enterprise", command)
        self.assertIn("--base-url", command)
        self.assertIn("https://github.fabrikam.example/api/v3", command)
        self.assertIn("--github-url", command)
        self.assertIn("FabrikamCloud", command)
        self.assertNotIn("--repo", command)
        self.assertIn("--store-lookup", command)
        self.assertIn("application_inventory_service", command)
        self.assertIn("--owner-user-id", command)
        self.assertIn("--owner-user-login", command)
        self.assertIn("--postgres-table", command)
        self.assertEqual(command.count("--application-type"), 3)
        self.assertIn("mobile_app", command)
        self.assertIn("ai_enabled", command)
        self.assertIn("ml_enabled", command)

    def test_normalize_scan_config_accepts_github_app_credentials(self):
        with patch.dict(
            "os.environ",
            {
                "APPLICATION_INVENTORY_GITHUB_APP_ID": "123",
                "APPLICATION_INVENTORY_GITHUB_APP_INSTALLATION_ID": "456",
            },
            clear=False,
        ):
            config = scanner.normalize_scan_config(
                {
                    "provider": "github-enterprise",
                    "org": "FabrikamCloud",
                    "githubAppPrivateKey": "-----BEGIN PRIVATE KEY-----\\nfake\\n-----END PRIVATE KEY-----",
                }
            )

        self.assertEqual(config["githubAppId"], "123")
        self.assertEqual(config["githubAppInstallationId"], "456")
        self.assertIn("BEGIN PRIVATE KEY", config["githubAppPrivateKey"])

        with patch.dict(
            "os.environ",
            {
                "APPLICATION_INVENTORY_GITHUB_APP_ID": "123",
                "APPLICATION_INVENTORY_GITHUB_APP_INSTALLATION_ID": "456",
            },
            clear=False,
        ):
            environment = ui_module.scan_environment(config)
        self.assertEqual(environment["APPLICATION_INVENTORY_GITHUB_APP_ID"], "123")
        self.assertEqual(
            environment["APPLICATION_INVENTORY_GITHUB_APP_INSTALLATION_ID"], "456"
        )
        self.assertIn(
            "BEGIN PRIVATE KEY",
            environment["APPLICATION_INVENTORY_GITHUB_APP_PRIVATE_KEY"],
        )

    def test_scan_environment_forwards_primary_github_app_private_key_file(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "github-enterprise",
                "org": "FabrikamCloud",
                "baseUrl": "https://api.github.com",
            }
        )

        with patch.dict(
            "os.environ",
            {
                "APPLICATION_INVENTORY_GITHUB_APP_PRIVATE_KEY_FILE": "/run/secrets/github-app.pem"
            },
            clear=False,
        ):
            environment = ui_module.scan_environment(config)

        self.assertEqual(
            environment["APPLICATION_INVENTORY_GITHUB_APP_PRIVATE_KEY_FILE"],
            "/run/secrets/github-app.pem",
        )

    def test_normalize_scan_config_accepts_mixed_sources(self):
        with patch.dict(
            "os.environ",
            {
                "APPLICATION_INVENTORY_GITHUB_APP_ID": "123",
                "APPLICATION_INVENTORY_GITHUB_APP_INSTALLATION_ID": "456",
            },
            clear=False,
        ):
            config = scanner.normalize_scan_config(
                {
                    "provider": "mixed",
                    "org": "FabrikamGH",
                    "adoOrgPats": [{"org": "FabrikamADO", "pat": "ado-token"}],
                    "githubAppPrivateKey": "-----BEGIN PRIVATE KEY-----\\nfake\\n-----END PRIVATE KEY-----",
                    "postgresEnabled": False,
                }
            )

        self.assertEqual(config["provider"], "mixed")
        self.assertEqual(config["orgDisplay"], "FabrikamGH + FabrikamADO")
        self.assertEqual(config["adoOrgPats"][0]["org"], "FabrikamADO")

    def test_build_scan_command_for_mixed_sources_uses_one_command(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "mixed",
                "org": "FabrikamGH",
                "baseUrl": "https://github.fabrikam.example/api/v3",
                "adoOrgPats": "FabrikamADO=ado-token",
                "postgresEnabled": False,
            }
        )

        command = scanner.build_scan_command(config, Path("/reports/scan-1"))
        environment = ui_module.scan_environment(config)

        self.assertEqual(command.count("--provider"), 1)
        self.assertIn("mixed", command)
        self.assertIn("--github-url", command)
        self.assertIn("FabrikamGH", command)
        self.assertIn("APPLICATION_INVENTORY_ADO_ORG_PATS", environment)
        self.assertNotIn("ado-token", command)

    def test_build_scan_command_for_selected_github_repositories(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "github-enterprise",
                "org": "FabrikamCloud",
                "baseUrl": "https://github.fabrikam.example/api/v3",
                "targetFilters": [
                    {"org": "FabrikamCloud", "project": "payments-api"},
                    {"org": "FabrikamCloud", "project": "storefront"},
                ],
            }
        )

        command = scanner.build_scan_command(config, Path("/reports/scan-1"))

        self.assertNotIn("--repo", command)
        self.assertEqual(command.count("--target-filter"), 2)
        self.assertIn("FabrikamCloud=payments-api", command)
        self.assertIn("FabrikamCloud=storefront", command)

    def test_normalize_scan_config_uses_fixed_prefix_and_postgres_default(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "azure-devops",
                "org": "FabrikamCloud",
                "adoOrgPats": "FabrikamCloud=pat-a",
                "outPrefix": "custom-prefix",
            }
        )

        self.assertEqual(config["outPrefix"], "application_inventory_service")
        self.assertTrue(config["postgresEnabled"])
        self.assertEqual(config["ownerUserId"], "anonymous")
        self.assertEqual(config["ownerUserLogin"], "anonymous")

    def test_normalize_scan_config_accepts_multiple_ado_org_pats_without_org(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "azure-devops",
                "adoOrgPats": "FabrikamCloud=pat-a\nContosoApps=pat-b",
                "postgresEnabled": False,
            }
        )

        self.assertEqual(config["org"], "")
        self.assertEqual(config["orgDisplay"], "2 Azure DevOps organizations")
        self.assertEqual(
            [(item["org"], item["pat"]) for item in config["adoOrgPats"]],
            [("FabrikamCloud", "pat-a"), ("ContosoApps", "pat-b")],
        )

    def test_normalize_scan_config_accepts_added_ado_org_pat_list(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "azure-devops",
                "adoOrgPats": [
                    {"org": "FabrikamCloud", "pat": "pat-a"},
                    {"org": "ContosoApps", "pat": "pat-b"},
                ],
                "postgresEnabled": False,
            }
        )

        self.assertEqual(config["orgDisplay"], "2 Azure DevOps organizations")
        self.assertEqual(
            [(item["org"], item["pat"]) for item in config["adoOrgPats"]],
            [("FabrikamCloud", "pat-a"), ("ContosoApps", "pat-b")],
        )

    def test_normalize_scan_config_accepts_selected_targets(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "azure-devops",
                "adoOrgPats": [{"org": "FabrikamCloud", "pat": "pat-a"}],
                "targetFilters": [
                    {"org": "FabrikamCloud", "project": "Payments"},
                    {"org": "FabrikamCloud", "project": "Storefront"},
                ],
                "postgresEnabled": False,
            }
        )

        self.assertEqual(
            [(item["org"], item["project"]) for item in config["targetFilters"]],
            [("FabrikamCloud", "Payments"), ("FabrikamCloud", "Storefront")],
        )

    def test_normalize_scan_config_rejects_unknown_application_types(self):
        with self.assertRaises(ValueError):
            scanner.normalize_scan_config(
                {
                    "provider": "azure-devops",
                    "org": "FabrikamCloud",
                    "adoOrgPats": "FabrikamCloud=pat-a",
                    "applicationTypes": ["mobile_app", "unknown"],
                }
            )

    def test_build_scan_command_scans_all_github_repos_when_repo_empty(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "github-enterprise",
                "org": "FabrikamCloud",
                "baseUrl": "https://github.fabrikam.example/api/v3",
            }
        )

        command = scanner.build_scan_command(config, Path("/reports/scan-1"))

        self.assertNotIn("--repo", command)

    def test_postgres_config_builds_environment_without_exposing_dsn_in_command(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "azure-devops",
                "org": "FabrikamCloud",
                "adoOrgPats": "FabrikamCloud=pat-a",
                "postgresEnabled": True,
                "postgresHost": "localhost",
                "postgresPort": 5432,
                "postgresDatabase": "postgres",
                "postgresUser": "postgres",
                "postgresPassword": "postgres",
                "postgresSchema": "application_inventory",
                "postgresTable": "application_inventory_assets",
            }
        )

        command = scanner.build_scan_command(config, Path("/reports/scan-1"))
        env = scanner.scan_environment(config)

        self.assertEqual(
            config["postgresDsn"],
            "postgresql://postgres:postgres@localhost:5432/postgres",
        )
        self.assertEqual(config["postgresSchema"], "application_inventory")
        self.assertIn("--postgres-schema", command)
        self.assertIn("application_inventory", command)
        self.assertIn("--postgres-table", command)
        self.assertIn("application_inventory_assets", command)
        self.assertNotIn(config["postgresDsn"], command)
        self.assertEqual(
            env["APPLICATION_INVENTORY_POSTGRES_DSN"], config["postgresDsn"]
        )
        self.assertEqual(
            env["APPLICATION_INVENTORY_POSTGRES_SCHEMA"], "application_inventory"
        )
        self.assertEqual(
            env["APPLICATION_INVENTORY_POSTGRES_TABLE"], "application_inventory_assets"
        )
        self.assertEqual(env["APPSEC_INVENTORY_POSTGRES_DSN"], config["postgresDsn"])
        self.assertEqual(
            env["APPSEC_INVENTORY_POSTGRES_SCHEMA"], "application_inventory"
        )
        self.assertEqual(
            env["APPSEC_INVENTORY_POSTGRES_TABLE"], "application_inventory_assets"
        )

    def test_postgres_disabled_removes_database_environment(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "azure-devops",
                "org": "FabrikamCloud",
                "adoOrgPats": "FabrikamCloud=pat-a",
                "postgresEnabled": False,
            }
        )

        with patch.dict(
            "os.environ",
            {
                "APPSEC_INVENTORY_POSTGRES_DSN": "postgresql://postgres:postgres@localhost:5432/postgres",
                "APPSEC_INVENTORY_POSTGRES_SCHEMA": "application_inventory",
                "APPSEC_INVENTORY_POSTGRES_TABLE": "application_inventory_assets",
                "APPLICATION_INVENTORY_POSTGRES_DSN": "postgresql://postgres:postgres@localhost:5432/postgres",
                "APPLICATION_INVENTORY_POSTGRES_SCHEMA": "application_inventory",
                "APPLICATION_INVENTORY_POSTGRES_TABLE": "application_inventory_assets",
            },
            clear=False,
        ):
            env = scanner.scan_environment(config)

        self.assertNotIn("APPSEC_INVENTORY_POSTGRES_DSN", env)
        self.assertNotIn("APPSEC_INVENTORY_POSTGRES_SCHEMA", env)
        self.assertNotIn("APPSEC_INVENTORY_POSTGRES_TABLE", env)
        self.assertNotIn("APPLICATION_INVENTORY_POSTGRES_DSN", env)
        self.assertNotIn("APPLICATION_INVENTORY_POSTGRES_SCHEMA", env)
        self.assertNotIn("APPLICATION_INVENTORY_POSTGRES_TABLE", env)

    def test_scan_environment_does_not_inherit_unrelated_secrets(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "azure-devops",
                "org": "FabrikamCloud",
                "adoOrgPats": "FabrikamCloud=pat-a",
                "token": "ado-token",
                "postgresEnabled": False,
            }
        )

        with patch.dict(
            "os.environ",
            {
                "AWS_SECRET_ACCESS_KEY": "aws-secret",
                "APPLICATION_INVENTORY_SERVICE_GHE_CLIENT_SECRET": "oauth-secret",
                "ADO_PAT": "environment-token",
            },
            clear=False,
        ):
            env = scanner.scan_environment(config)

        self.assertNotIn("ADO_PAT", env)
        self.assertIn("FabrikamCloud", env["APPLICATION_INVENTORY_ADO_ORG_PATS"])
        self.assertNotIn("AWS_SECRET_ACCESS_KEY", env)
        self.assertNotIn("APPLICATION_INVENTORY_SERVICE_GHE_CLIENT_SECRET", env)

    def test_multi_ado_org_pats_are_passed_through_environment(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "azure-devops",
                "adoOrgPats": "FabrikamCloud=pat-a\nContosoApps=pat-b",
                "postgresEnabled": False,
            }
        )

        command = scanner.build_scan_command(config, Path("/reports/scan-1"))
        env = scanner.scan_environment(config)

        self.assertNotIn("--org", command)
        self.assertNotIn("pat-a", command)
        self.assertNotIn("pat-b", command)
        self.assertIn("APPLICATION_INVENTORY_ADO_ORG_PATS", env)
        self.assertIn("APPSEC_INVENTORY_ADO_ORG_PATS", env)
        self.assertNotIn("ADO_PAT", env)
        self.assertEqual(
            json.loads(env["APPLICATION_INVENTORY_ADO_ORG_PATS"]),
            [
                {"org": "FabrikamCloud", "pat": "pat-a"},
                {"org": "ContosoApps", "pat": "pat-b"},
            ],
        )
        self.assertEqual(
            json.loads(env["APPSEC_INVENTORY_ADO_ORG_PATS"]),
            [
                {"org": "FabrikamCloud", "pat": "pat-a"},
                {"org": "ContosoApps", "pat": "pat-b"},
            ],
        )

    def test_postgres_dsn_url_encodes_credentials(self):
        dsn = scanner.postgres_dsn_from_config(
            {
                "postgresHost": "localhost",
                "postgresPort": 5432,
                "postgresDatabase": "inventory db",
                "postgresUser": "app user",
                "postgresPassword": "p@ss word",
            }
        )

        self.assertEqual(
            dsn, "postgresql://app%20user:p%40ss%20word@localhost:5432/inventory%20db"
        )

    def test_database_status_requires_dsn(self):
        status = scanner.database_status("")

        self.assertFalse(status["connected"])
        self.assertEqual(status["status"], "missing_dsn")

    def test_postgres_inventory_key_is_user_scoped(self):
        self.assertEqual(
            PRIMARY_KEY_COLUMNS,
            (
                "owner_user_id",
                "provider",
                "organization",
                "project",
                "repo_name",
                "branch_name",
            ),
        )

    def test_database_search_normalizes_and_bounds_terms(self):
        query = "  mobile\n app   " + " ".join(f"term-{index}" for index in range(20))

        normalized = normalize_search_query(query)
        tokens = search_tokens(query)

        self.assertFalse(normalized.startswith(" "))
        self.assertNotIn("\n", normalized)
        self.assertEqual(tokens[:2], ["mobile", "app"])
        self.assertEqual(len(tokens), 12)

    def test_build_scan_command_for_azure_devops(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "azure-devops",
                "org": "FabrikamCloud",
                "adoOrgPats": "FabrikamCloud=pat-a",
                "project": "Go_To_Market",
                "activityMode": "contributors",
                "maxWorkers": 4,
            }
        )

        command = scanner.build_scan_command(config, Path("/reports/scan-2"))

        self.assertIn("--provider", command)
        self.assertIn("azure-devops", command)
        self.assertIn("--project", command)
        self.assertIn("Go_To_Market", command)
        self.assertNotIn("--base-url", command)
        self.assertIn("4", command)

    def test_build_scan_command_for_selected_ado_projects(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "azure-devops",
                "adoOrgPats": [{"org": "FabrikamCloud", "pat": "pat-a"}],
                "targetFilters": [{"org": "FabrikamCloud", "project": "Payments"}],
                "postgresEnabled": False,
            }
        )

        command = scanner.build_scan_command(config, Path("/reports/scan-2"))

        self.assertNotIn("--project", command)
        self.assertEqual(command.count("--target-filter"), 1)
        self.assertIn("FabrikamCloud=Payments", command)

    def test_build_scan_command_scans_all_ado_projects_when_project_empty(self):
        config = scanner.normalize_scan_config(
            {
                "provider": "azure-devops",
                "org": "FabrikamCloud",
                "adoOrgPats": "FabrikamCloud=pat-a",
            }
        )

        command = scanner.build_scan_command(config, Path("/reports/scan-2"))

        self.assertNotIn("--project", command)

    def test_discover_source_targets_loads_azure_projects(self):
        created = []

        class FakeAzureClient:
            def __init__(self, org, pat, timeout):
                created.append((org, pat, timeout))

            def list_projects(self):
                return [{"name": "Payments"}, {"name": "Storefront"}]

            def close(self):
                return None

        with patch.object(
            source_discovery_module, "AzureDevOpsClient", FakeAzureClient
        ):
            data = ui_module.discover_source_targets(
                {
                    "provider": "azure-devops",
                    "adoOrgPats": [{"org": "FabrikamCloud", "pat": "pat-a"}],
                    "timeout": 12,
                }
            )

        self.assertEqual(created, [("FabrikamCloud", "pat-a", 12)])
        self.assertEqual(
            [
                (target["org"], target["project"], target["kind"])
                for target in data["targets"]
            ],
            [
                ("FabrikamCloud", "Payments", "project"),
                ("FabrikamCloud", "Storefront", "project"),
            ],
        )

    def test_discover_source_targets_loads_github_repositories(self):
        created = []

        class FakeGitHubClient:
            def __init__(self, base_url, owner, token, timeout, **kwargs):
                created.append((base_url, owner, token, timeout))

            def list_repos(self, project_name):
                if project_name != "":
                    raise AssertionError(
                        "repository discovery should list all repositories"
                    )
                return [
                    {"name": "payments-api", "fullName": "FabrikamCloud/payments-api"},
                    {
                        "name": "archived-api",
                        "fullName": "FabrikamCloud/archived-api",
                        "isDisabled": True,
                    },
                ]

            def close(self):
                return None

        with (
            patch.object(
                source_discovery_module, "GitHubEnterpriseClient", FakeGitHubClient
            ),
            patch.dict(
                "os.environ",
                {
                    "APPLICATION_INVENTORY_GITHUB_API_URL": "https://github.fabrikam.example/api/v3"
                },
                clear=False,
            ),
        ):
            data = ui_module.discover_source_targets(
                {
                    "provider": "github-enterprise",
                    "org": "FabrikamCloud",
                    "token": "token",
                    "timeout": 11,
                }
            )

        self.assertEqual(
            created,
            [("https://github.fabrikam.example/api/v3", "FabrikamCloud", "token", 11)],
        )
        self.assertEqual(
            [
                (target["org"], target["project"], target["kind"], target["label"])
                for target in data["targets"]
            ],
            [
                (
                    "FabrikamCloud",
                    "payments-api",
                    "repository",
                    "FabrikamCloud/payments-api",
                )
            ],
        )

    def test_normalize_application_types_uses_known_order(self):
        self.assertEqual(
            scanner.normalize_application_types(
                ["ml_enabled", "ai_enabled", "mobile_app", "mobile_app"]
            ),
            ("mobile_app", "ai_enabled", "ml_enabled"),
        )

    def test_inventory_type_matches_defaults_to_all_types(self):
        self.assertTrue(scanner.inventory_type_matches(["web_app"], ()))
        self.assertTrue(
            scanner.inventory_type_matches(["web_app", "api_service"], ["api_service"])
        )
        self.assertFalse(scanner.inventory_type_matches(["web_app"], ["mobile_app"]))

    def test_store_lookup_is_mobile_only(self):
        self.assertTrue(scanner.store_lookup_allowed(()))
        self.assertTrue(scanner.store_lookup_allowed(("mobile_app",)))
        self.assertFalse(scanner.store_lookup_allowed(("web_app", "ai_enabled")))
        config = scanner.normalize_scan_config(
            {
                "provider": "azure-devops",
                "org": "FabrikamCloud",
                "adoOrgPats": "FabrikamCloud=pat-a",
                "applicationTypes": ["web_app"],
                "storeLookup": True,
            }
        )
        self.assertFalse(config["storeLookup"])

    def test_redact_command_hides_sensitive_values(self):
        command = (
            "application-inventory-service",
            "--pat",
            "secret",
            "--postgres-dsn",
            "postgresql://postgres:postgres@localhost:5432/postgres",
            "--org",
            "FabrikamCloud",
        )

        redacted = scanner.redact_command(command)

        self.assertEqual(
            redacted,
            [
                "application-inventory-service",
                "--pat",
                "[redacted]",
                "--postgres-dsn",
                "[redacted]",
                "--org",
                "FabrikamCloud",
            ],
        )


class DetectionTests(unittest.TestCase):
    def test_detects_react_native_android_repo(self):
        paths = [
            "/package.json",
            "/android/app/build.gradle",
            "/android/app/src/main/AndroidManifest.xml",
            "/metro.config.js",
        ]
        contents = {
            "/package.json": '{"dependencies": {"react-native": "0.75.0"}}',
            "/android/app/build.gradle": """\
plugins {
    id 'com.android.application'
}

android {
    namespace 'com.fabrikam.agsnap'
    defaultConfig {
        applicationId 'com.fabrikam.agsnap'
        versionName '1.0.2'
    }
}
""",
            "/android/app/src/main/AndroidManifest.xml": """\
<manifest xmlns:android="http://schemas.android.com/apk/res/android"
    package="com.fabrikam.agsnap" />
""",
        }

        confidence, evidence, score = scanner.detect_mobile_repo(paths, contents)
        categories = {item.category for item in evidence}

        self.assertEqual(confidence, "high")
        self.assertGreaterEqual(score, 7)
        self.assertIn("android", categories)
        self.assertIn("react_native", categories)

    def test_returns_none_when_no_mobile_signals_exist(self):
        confidence, evidence, score = scanner.detect_mobile_repo(
            ["/README.md", "/src/server.py"],
            {},
        )

        self.assertEqual(confidence, "none")
        self.assertEqual(evidence, [])
        self.assertEqual(score, 0)

    def test_detects_web_frontend_inventory_repo(self):
        confidence, evidence, score = scanner.detect_inventory_repo(
            ["/package.json", "/src/App.tsx"],
            {
                "/package.json": json.dumps(
                    {
                        "name": "customer-portal",
                        "version": "2.4.1",
                        "dependencies": {"react": "18.3.1", "vite": "5.0.0"},
                        "scripts": {"start": "vite"},
                    }
                )
            },
        )

        categories = {item.category for item in evidence}
        self.assertEqual(confidence, "medium")
        self.assertGreaterEqual(score, 4)
        self.assertIn("web_frontend", categories)

    def test_detects_spring_microservice_inventory_repo(self):
        confidence, evidence, score = scanner.detect_inventory_repo(
            ["/pom.xml", "/Dockerfile"],
            {
                "/pom.xml": """\
<project>
  <artifactId>orders-api</artifactId>
  <dependencies>
    <dependency><artifactId>spring-boot-starter-web</artifactId></dependency>
    <dependency><artifactId>spring-kafka</artifactId></dependency>
  </dependencies>
</project>
""",
                "/Dockerfile": 'FROM eclipse-temurin:21\nCMD ["java", "-jar", "app.jar"]\n',
            },
        )

        categories = {item.category for item in evidence}
        self.assertEqual(confidence, "high")
        self.assertGreaterEqual(score, 8)
        self.assertIn("microservice", categories)
        self.assertIn("api_service", categories)
        self.assertIn("middleware", categories)
        self.assertIn("containerized_service", categories)

    def test_detects_python_middleware_inventory_repo(self):
        confidence, evidence, score = scanner.detect_inventory_repo(
            ["/pyproject.toml"],
            {
                "/pyproject.toml": """\
[project]
name = "billing-worker"
version = "0.8.0"
dependencies = ["celery", "confluent-kafka"]
"""
            },
        )

        categories = {item.category for item in evidence}
        self.assertEqual(confidence, "medium")
        self.assertGreaterEqual(score, 3)
        self.assertIn("middleware", categories)

    def test_detects_node_llm_inventory_repo(self):
        confidence, evidence, score = scanner.detect_inventory_repo(
            ["/package.json"],
            {
                "/package.json": json.dumps(
                    {
                        "name": "support-copilot",
                        "version": "1.7.0",
                        "dependencies": {
                            "@anthropic-ai/sdk": "0.56.0",
                            "@langchain/openai": "0.5.0",
                            "@pinecone-database/pinecone": "5.1.0",
                        },
                    }
                )
            },
        )

        categories = {item.category for item in evidence}
        self.assertEqual(confidence, "high")
        self.assertGreaterEqual(score, 8)
        self.assertIn("ai_enabled", categories)
        self.assertIn("llm_integration", categories)
        self.assertIn("ai_orchestration", categories)
        self.assertIn("vector_search", categories)

    def test_detects_python_ai_inventory_repo(self):
        confidence, evidence, score = scanner.detect_inventory_repo(
            ["/pyproject.toml"],
            {
                "/pyproject.toml": """\
[project]
name = "claims-ai"
version = "0.3.0"
dependencies = ["openai", "langchain-openai", "chromadb", "sentence-transformers"]
"""
            },
        )

        categories = {item.category for item in evidence}
        self.assertEqual(confidence, "high")
        self.assertGreaterEqual(score, 8)
        self.assertIn("ai_enabled", categories)
        self.assertIn("llm_integration", categories)
        self.assertIn("ai_orchestration", categories)
        self.assertIn("ml_inference", categories)
        self.assertIn("ml_enabled", categories)
        self.assertIn("vector_search", categories)

    def test_detects_ml_enabled_inventory_repo(self):
        confidence, evidence, score = scanner.detect_inventory_repo(
            ["/requirements.txt"],
            {"/requirements.txt": "fastapi\nscikit-learn\nmlflow\nonnxruntime\n"},
        )

        categories = {item.category for item in evidence}
        inventory_types = scanner.inventory_types_from_categories(categories)

        self.assertEqual(confidence, "high")
        self.assertGreaterEqual(score, 8)
        self.assertIn("ai_enabled", categories)
        self.assertIn("ml_enabled", categories)
        self.assertIn("ml_inference", categories)
        self.assertIn("ai_enabled", inventory_types)
        self.assertIn("ml_enabled", inventory_types)

    def test_detects_dotnet_ai_inventory_repo(self):
        confidence, evidence, score = scanner.detect_inventory_repo(
            ["/Assistant.csproj"],
            {
                "/Assistant.csproj": """\
<Project Sdk="Microsoft.NET.Sdk.Web">
  <ItemGroup>
    <PackageReference Include="Azure.AI.OpenAI" Version="2.1.0" />
    <PackageReference Include="Microsoft.SemanticKernel" Version="1.55.0" />
  </ItemGroup>
</Project>
"""
            },
        )

        categories = {item.category for item in evidence}
        self.assertEqual(confidence, "high")
        self.assertGreaterEqual(score, 8)
        self.assertIn("web_backend", categories)
        self.assertIn("api_service", categories)
        self.assertIn("ai_enabled", categories)
        self.assertIn("llm_integration", categories)
        self.assertIn("ai_orchestration", categories)

    def test_detects_ai_runtime_configuration(self):
        confidence, evidence, score = scanner.detect_inventory_repo(
            ["/src/main/resources/application.yml"],
            {
                "/src/main/resources/application.yml": """\
spring:
  application:
    name: assistant-api
  ai:
    openai:
      chat:
        options:
          model: gpt-4.1-mini
"""
            },
        )

        categories = {item.category for item in evidence}
        self.assertEqual(confidence, "medium")
        self.assertGreaterEqual(score, 3)
        self.assertIn("ai_enabled", categories)
        self.assertIn("llm_integration", categories)

    def test_detects_ml_runtime_configuration(self):
        confidence, evidence, score = scanner.detect_inventory_repo(
            ["/application.yml"],
            {
                "/application.yml": """\
mlflow:
  tracking_uri: https://mlflow.example.invalid
model_uri: models:/churn/latest
"""
            },
        )

        categories = {item.category for item in evidence}
        self.assertEqual(confidence, "medium")
        self.assertGreaterEqual(score, 7)
        self.assertIn("ai_enabled", categories)
        self.assertIn("ml_enabled", categories)
        self.assertIn("ml_inference", categories)

    def test_generic_config_xml_is_not_mobile(self):
        confidence, evidence, score = scanner.detect_mobile_repo(
            ["/config.xml"],
            {
                "/config.xml": "<configuration><setting name='example' /></configuration>"
            },
        )

        self.assertEqual(confidence, "none")
        self.assertEqual(evidence, [])
        self.assertEqual(score, 0)

    def test_generic_csproj_is_not_xamarin_or_maui(self):
        confidence, evidence, score = scanner.detect_mobile_repo(
            ["/src/Api/Api.csproj"],
            {
                "/src/Api/Api.csproj": """\
<Project Sdk="Microsoft.NET.Sdk.Web">
  <PropertyGroup>
    <TargetFramework>net8.0</TargetFramework>
  </PropertyGroup>
</Project>
"""
            },
        )

        self.assertEqual(confidence, "none")
        self.assertEqual(evidence, [])
        self.assertEqual(score, 0)

    def test_resolved_gradle_app_id_contributes_to_detection_evidence(self):
        confidence, evidence, score = scanner.detect_mobile_repo(
            ["/gradle.properties", "/android/app/build.gradle"],
            {
                "/gradle.properties": "appId=com.fabrikam.agsnap\n",
                "/android/app/build.gradle": """\
plugins {
    id 'com.android.application'
}
android {
    defaultConfig {
        applicationId "${appId}"
    }
}
""",
            },
        )

        details = {item.detail for item in evidence}
        self.assertEqual(confidence, "high")
        self.assertGreaterEqual(score, 7)
        self.assertIn("Gradle applicationId com.fabrikam.agsnap", details)

    def test_should_fetch_allowed_content_files(self):
        self.assertTrue(scanner.should_fetch_content("/src/MyApp.csproj"))
        self.assertTrue(scanner.should_fetch_content("/package.json"))
        self.assertTrue(scanner.should_fetch_content("/pom.xml"))
        self.assertTrue(scanner.should_fetch_content("/pyproject.toml"))
        self.assertTrue(scanner.should_fetch_content("/Dockerfile"))
        self.assertTrue(scanner.should_fetch_content("/serverless.yml"))
        self.assertTrue(scanner.should_fetch_content("/gradle.properties"))
        self.assertTrue(scanner.should_fetch_content("/Directory.Build.props"))
        self.assertTrue(
            scanner.should_fetch_content("/android/app/src/main/AndroidManifest.xml")
        )
        self.assertTrue(scanner.should_fetch_content("/ios/App/Info.plist"))
        self.assertTrue(
            scanner.should_fetch_content("/android/app/src/main/res/values/strings.xml")
        )
        self.assertFalse(scanner.should_fetch_content("/src/app.py"))

    def test_normalize_path_adds_leading_slash_and_unix_separators(self):
        self.assertEqual(
            scanner.normalize_path("android\\app\\build.gradle"),
            "/android/app/build.gradle",
        )


class MetadataExtractionTests(unittest.TestCase):
    def test_extracts_android_name_version_and_identifier(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/android/app/src/main/AndroidManifest.xml": """\
<manifest xmlns:android="http://schemas.android.com/apk/res/android"
    package="com.fabrikam.agsnap"
    android:versionName="1.0.2">
    <application android:label="@string/app_name" />
</manifest>
""",
                "/android/app/src/main/res/values/strings.xml": """\
<resources>
    <string name="app_name">Agsnap</string>
</resources>
""",
            }
        )

        self.assertEqual(metadata.name, "Agsnap")
        self.assertEqual(metadata.version, "1.0.2")
        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")

    def test_extracts_gradle_identifier_and_version(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/android/app/build.gradle.kts": """\
android {
    namespace = "com.fabrikam.agsnap"
    defaultConfig {
        applicationId = "com.fabrikam.agsnap"
        versionName = "1.0.2"
    }
}
"""
            }
        )

        self.assertEqual(metadata.version, "1.0.2")
        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")
        self.assertEqual(metadata.identifier_source, "Gradle applicationId/namespace")

    def test_resolves_gradle_identifier_from_properties(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/gradle.properties": "appId=com.fabrikam.agsnap\n",
                "/android/app/build.gradle": """\
android {
    defaultConfig {
        applicationId "${appId}"
    }
}
""",
            }
        )

        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")
        self.assertEqual(metadata.identifier_source, "Gradle applicationId/namespace")

    def test_extracts_ios_info_plist_metadata(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/ios/App/Info.plist": """\
<?xml version="1.0" encoding="UTF-8"?>
<plist version="1.0">
<dict>
  <key>CFBundleDisplayName</key><string>Agsnap</string>
  <key>CFBundleShortVersionString</key><string>1.0.2</string>
  <key>CFBundleIdentifier</key><string>com.fabrikam.agsnap</string>
</dict>
</plist>
"""
            }
        )

        self.assertEqual(metadata.name, "Agsnap")
        self.assertEqual(metadata.version, "1.0.2")
        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")

    def test_extracts_ios_metadata_from_xcode_settings(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/ios/App.xcodeproj/project.pbxproj": """\
PRODUCT_NAME = Agsnap;
MARKETING_VERSION = 1.0.2;
PRODUCT_BUNDLE_IDENTIFIER = com.fabrikam.agsnap;
"""
            }
        )

        self.assertEqual(metadata.name, "Agsnap")
        self.assertEqual(metadata.version, "1.0.2")
        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")
        self.assertEqual(metadata.identifier_source, "Xcode build settings")

    def test_resolves_ios_plist_identifier_from_xcode_settings(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/ios/App/Info.plist": """\
<plist version="1.0">
<dict>
  <key>CFBundleIdentifier</key><string>$(PRODUCT_BUNDLE_IDENTIFIER)</string>
</dict>
</plist>
""",
                "/ios/App.xcodeproj/project.pbxproj": """\
PRODUCT_BUNDLE_IDENTIFIER = $(PRODUCT_BUNDLE_IDENTIFIER);
PRODUCT_BUNDLE_IDENTIFIER = com.fabrikam.agsnap;
""",
            }
        )

        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")
        self.assertEqual(metadata.identifier_source, "Info.plist")

    def test_filters_placeholder_versions(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/package.json": json.dumps(
                    {
                        "name": "agsnap",
                        "version": "999.999.999",
                    }
                )
            }
        )

        self.assertEqual(metadata.name, "agsnap")
        self.assertEqual(metadata.version, "")

    def test_extracts_expo_metadata(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/app.json": json.dumps(
                    {
                        "expo": {
                            "name": "Agsnap",
                            "version": "1.0.2",
                            "ios": {"bundleIdentifier": "com.fabrikam.agsnap"},
                            "android": {"package": "com.fabrikam.agsnap"},
                        }
                    }
                )
            }
        )

        self.assertEqual(metadata.name, "Agsnap")
        self.assertEqual(metadata.version, "1.0.2")
        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")

    def test_extracts_capacitor_metadata(self):
        metadata = scanner.extract_mobile_metadata(
            {
                "/capacitor.config.ts": """\
export default {
  appId: 'com.fabrikam.agsnap',
  appName: 'Agsnap',
  version: '1.0.2'
}
"""
            }
        )

        self.assertEqual(metadata.name, "Agsnap")
        self.assertEqual(metadata.version, "1.0.2")
        self.assertEqual(metadata.identifier, "com.fabrikam.agsnap")


class RepoActivityTests(unittest.TestCase):
    def test_extracts_contributors_and_last_updated(self):
        activity = scanner.extract_repo_activity(
            [
                {
                    "author": {"name": "Alice Adams", "email": "alice@example.com"},
                    "committer": {
                        "name": "Build Service",
                        "email": "build@example.com",
                        "date": "2024-04-01T12:00:00Z",
                    },
                },
                {
                    "author": {"name": "Bob Brown", "email": "bob@example.com"},
                    "committer": {
                        "name": "Bob Brown",
                        "email": "bob@example.com",
                        "date": "2024-05-02T08:30:15.123Z",
                    },
                },
                {
                    "author": {"name": "Alice Adams", "email": "alice@example.com"},
                    "committer": {
                        "name": "Alice Adams",
                        "email": "alice@example.com",
                        "date": "2024-03-01T00:00:00Z",
                    },
                },
            ]
        )

        self.assertEqual(
            activity.contributing_developers,
            (
                "Alice Adams <alice@example.com>",
                "Bob Brown <bob@example.com>",
            ),
        )
        self.assertEqual(activity.last_updated, "2024-05-02T08:30:15Z")

    def test_fetch_repo_activity_latest_mode_only_requests_latest_commit(self):
        class FakeClient:
            def __init__(self):
                self.calls = []

            def list_commits(self, **kwargs):
                self.calls.append(kwargs)
                return [
                    {
                        "author": {"name": "Alice Adams", "email": "alice@example.com"},
                        "committer": {"date": "2024-05-02T08:30:15Z"},
                    }
                ]

        client = FakeClient()

        activity = scanner.fetch_repo_activity(
            client=client,
            project_name="Project",
            repo_id="repo-id",
            branch_name="main",
            max_commits=0,
            activity_mode="latest",
        )

        self.assertEqual(client.calls[0]["max_commits"], 1)
        self.assertEqual(activity.contributing_developers, ())
        self.assertEqual(activity.last_updated, "2024-05-02T08:30:15Z")

    def test_fetch_repo_activity_contributors_mode_uses_requested_commit_limit(self):
        class FakeClient:
            def __init__(self):
                self.calls = []

            def list_commits(self, **kwargs):
                self.calls.append(kwargs)
                return []

        client = FakeClient()

        scanner.fetch_repo_activity(
            client=client,
            project_name="Project",
            repo_id="repo-id",
            branch_name="main",
            max_commits=250,
            activity_mode="contributors",
        )

        self.assertEqual(client.calls[0]["max_commits"], 250)


class OutputTests(unittest.TestCase):
    class FakeBranchClient:
        def __init__(self, refs=None, definitions=None):
            self.refs = refs or []
            self.definitions = definitions or []
            self.branch_calls = 0
            self.definition_calls = 0

        def list_branches(self, project_name, repo_id):
            self.branch_calls += 1
            return self.refs

        def list_build_definitions_for_repo(self, project_name, repo_id):
            self.definition_calls += 1
            return self.definitions

    def sample_result(self):
        return {
            "provider": "azure-devops",
            "organization": "FabrikamCloud",
            "project": "Project",
            "repo_name": "Repo",
            "branch_name": "main",
            "branch_last_updated": "2024-05-02T08:30:15Z",
            "branch_age_bucket": scanner.ACTIVE_SHEET_NAME,
            "web_url": "https://example.invalid/repo",
            "source_url": "https://example.invalid/repo.git",
            "primary_web_domain": "agsnap.fabrikam.example",
            "web_domains": "agsnap.fabrikam.example",
            "web_urls": "https://agsnap.fabrikam.example",
            "web_domain_status": "configured",
            "web_domain_sources": "agsnap.fabrikam.example [source:/deploy/ingress.yaml:host:3]",
            "web_domain_evidence": json.dumps(
                [
                    {
                        "domain": "agsnap.fabrikam.example",
                        "url": "https://agsnap.fabrikam.example",
                        "confidence": "configured",
                        "environment": "production",
                        "sources": ["source:/deploy/ingress.yaml:host:3"],
                    }
                ]
            ),
            "inventory_name": "Agsnap",
            "inventory_version": "1.0.2",
            "inventory_types": "mobile_app",
            "primary_language": "Java/Kotlin",
            "scanner_target": "https://example.invalid/repo.git#branch=main",
            "semgrep_target": "https://example.invalid/repo.git#branch=main",
            "sonarqube_project_key": "Project:Repo:main",
            "sonarqube_project_name": "Agsnap",
            "mobile_name": "Agsnap",
            "mobile_version": "1.0.2",
            "mobile_identifier": "com.fabrikam.agsnap",
            "mobile_identifier_source": "Gradle applicationId/namespace",
            "mobile_identifier_status": "found",
            "nowsecure_target": "https://example.invalid/repo.git#branch=main",
            "branch_contributing_developers": "Alice Adams <alice@example.com>; Bob Brown <bob@example.com>",
            "contributing_developers": "Alice Adams <alice@example.com>; Bob Brown <bob@example.com>",
            "last_updated": "2024-05-02T08:30:15Z",
            "confidence": "medium",
            "score": 2,
            "categories": "android",
            **scanner.type_columns(["mobile_app"]),
            **scanner.category_columns(["android"]),
            "detection_evidence": json.dumps(
                [
                    {
                        "category": "android",
                        "source": "/android/app/build.gradle",
                        "detail": "Gradle applicationId com.fabrikam.agsnap",
                        "weight": 3,
                    }
                ]
            ),
        }

    def test_write_outputs_creates_labeled_scanner_outputs(self):
        result = self.sample_result()

        with tempfile.TemporaryDirectory() as tmpdir:
            xlsx_path, semgrep_path, sonarqube_path = scanner.write_outputs(
                [result],
                Path(tmpdir),
                "scan",
                application_types=("mobile_app",),
            )

            self.assertTrue(xlsx_path.exists())
            self.assertTrue(semgrep_path.exists())
            self.assertTrue(sonarqube_path.exists())
            self.assertEqual(xlsx_path.name, "scan_mobile_app.xlsx")
            self.assertEqual(semgrep_path.name, "scan_mobile_app_semgrep_targets.txt")
            self.assertEqual(
                sonarqube_path.name, "scan_mobile_app_sonarqube_projects.csv"
            )
            self.assertFalse((Path(tmpdir) / "scan_mobile_app.csv").exists())
            self.assertFalse((Path(tmpdir) / "scan_mobile_app.json").exists())
            self.assertFalse(
                (Path(tmpdir) / "scan_mobile_app_scanner_targets.csv").exists()
            )
            self.assertFalse(
                (Path(tmpdir) / "scan_mobile_app_scanner_targets.json").exists()
            )
            self.assertIn(
                "https://example.invalid/repo.git#branch=main",
                semgrep_path.read_text(encoding="utf-8"),
            )
            sonar_text = sonarqube_path.read_text(encoding="utf-8")
            self.assertIn("sonar.projectKey", sonar_text)
            self.assertIn("Agsnap", sonar_text)
            workbook = load_workbook(xlsx_path)
            self.assertEqual(
                workbook.sheetnames,
                [scanner.ACTIVE_SHEET_NAME, scanner.OLDER_SHEET_NAME],
            )
            self.assertEqual(
                workbook_value(workbook[scanner.ACTIVE_SHEET_NAME], "mobile_name", 2),
                "Agsnap",
            )
            self.assertEqual(
                workbook_value(workbook[scanner.ACTIVE_SHEET_NAME], "provider", 2),
                "azure-devops",
            )
            self.assertEqual(
                workbook_value(
                    workbook[scanner.ACTIVE_SHEET_NAME], "primary_web_domain", 2
                ),
                "agsnap.fabrikam.example",
            )
            self.assertEqual(
                workbook_value(
                    workbook[scanner.ACTIVE_SHEET_NAME],
                    "branch_contributing_developers",
                    2,
                ),
                "Alice Adams <alice@example.com>; Bob Brown <bob@example.com>",
            )
            self.assertEqual(
                workbook_value(
                    workbook[scanner.ACTIVE_SHEET_NAME], "nowsecure_target", 2
                ),
                "https://example.invalid/repo.git#branch=main",
            )

    def test_mobile_routing_and_store_metadata_only_apply_to_mobile_apps(self):
        class FailingStoreClient:
            def lookup(self, identifier, categories):
                raise AssertionError("non-mobile applications must not query stores")

        target = scanner.RepoScanTarget(
            project_name="Project",
            repo={
                "name": "service",
                "remoteUrl": "https://example.invalid/service.git",
            },
        )
        common = {
            "target": target,
            "branch_name": "main",
            "contents": {},
            "paths": [],
            "activity": scanner.RepoActivityMetadata(),
            "confidence": "high",
            "score": 3,
            "evidence": [],
            "branch_age_days": 90,
        }

        service = scanner_module.build_scan_row(
            **common,
            metadata=scanner.MobileAppMetadata(identifier="com.fabrikam.service"),
            categories=["microservice"],
            store_client=FailingStoreClient(),
        )
        mobile = scanner_module.build_scan_row(
            **common,
            metadata=scanner.MobileAppMetadata(identifier="com.fabrikam.mobile"),
            categories=["android"],
            store_client=None,
        )

        self.assertEqual(service["nowsecure_target"], "")
        self.assertNotIn("store_validation_passed", service)
        self.assertEqual(
            mobile["nowsecure_target"],
            "https://example.invalid/service.git#branch=main",
        )
        self.assertEqual(mobile["store_lookup_status"], "disabled")

    def test_cli_report_path_streams_without_retaining_result_rows(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            config = scanner.ScanConfig(
                org="FabrikamCloud",
                pat="token",
                project=None,
                out_dir=Path(tmpdir),
                out_prefix="scan",
                max_workers=1,
                content_workers=1,
                max_commits_per_repo=1,
                timeout_seconds=30,
                min_confidence="medium",
            )
            rows = [self.sample_result(), self.sample_result()]

            def fake_scan(scan_config, on_result=None, retain_results=True):
                self.assertFalse(retain_results)
                for row in rows:
                    on_result(row)
                return []

            with patch.object(scanner_module, "scan", side_effect=fake_scan):
                result_count, xlsx_path, _, _ = scanner.scan_reports(config)

            workbook = load_workbook(xlsx_path)
            max_row = workbook[scanner.ACTIVE_SHEET_NAME].max_row
            workbook.close()

        self.assertEqual(result_count, 2)
        self.assertEqual(max_row, 3)

    def test_streaming_report_writer_flushes_rows_as_they_are_written(self):
        result = self.sample_result()

        with tempfile.TemporaryDirectory() as tmpdir:
            with scanner.StreamingReportWriter(Path(tmpdir), "scan") as writer:
                self.assertTrue(writer.xlsx_path.exists())
                self.assertTrue(writer.semgrep_targets_path.exists())
                self.assertTrue(writer.sonarqube_projects_path.exists())

                writer.write_result(result)

                semgrep_text = writer.semgrep_targets_path.read_text(encoding="utf-8")
                sonarqube_text = writer.sonarqube_projects_path.read_text(
                    encoding="utf-8"
                )
                self.assertIn(
                    "https://example.invalid/repo.git#branch=main", semgrep_text
                )
                self.assertIn("Agsnap", sonarqube_text)

            workbook = load_workbook(writer.xlsx_path)
            self.assertEqual(
                workbook_value(workbook[scanner.ACTIVE_SHEET_NAME], "mobile_name", 2),
                "Agsnap",
            )

    def test_workbook_checkpoints_expand_to_avoid_repeated_full_serialization(self):
        settings = {
            "APPLICATION_INVENTORY_XLSX_CHECKPOINT_ROWS": "10",
            "APPLICATION_INVENTORY_XLSX_MAX_CHECKPOINT_ROWS": "40",
            "APPLICATION_INVENTORY_XLSX_CHECKPOINT_SECONDS": "3600",
        }
        with (
            tempfile.TemporaryDirectory() as tmpdir,
            patch.dict("os.environ", settings),
        ):
            with scanner.StreamingReportWriter(Path(tmpdir), "scan") as writer:
                with (
                    patch.object(writer, "_save_workbook") as save_workbook,
                    patch.object(writer, "_append_workbook_row"),
                    patch.object(writer, "flush"),
                ):
                    for _ in range(100):
                        writer.write_result({})

                self.assertEqual(save_workbook.call_count, 4)

    def test_streaming_report_writer_removes_illegal_workbook_characters(self):
        result = self.sample_result()
        result["contributing_developers"] = "Alice\x08 Adams <alice@example.com>"

        with tempfile.TemporaryDirectory() as tmpdir:
            with scanner.StreamingReportWriter(Path(tmpdir), "scan") as writer:
                writer.write_result(result)

            workbook = load_workbook(writer.xlsx_path)
            self.assertEqual(
                workbook_value(
                    workbook[scanner.ACTIVE_SHEET_NAME], "contributing_developers", 2
                ),
                "Alice Adams <alice@example.com>",
            )

    def test_workbook_cell_value_removes_illegal_characters(self):
        self.assertEqual(scanner.workbook_cell_value("Ag\x08snap"), "Agsnap")
        self.assertEqual(scanner.workbook_cell_value(42), 42)

    def test_postgres_rows_include_owner_scope(self):
        config = scanner.ScanConfig(
            org="FabrikamCloud",
            pat="token",
            project=None,
            out_dir=Path("reports"),
            out_prefix="scan",
            max_workers=1,
            branch_workers=1,
            content_workers=1,
            max_commits_per_repo=0,
            timeout_seconds=30,
            min_confidence="low",
            owner_user_id="42",
            owner_user_login="alice",
        )
        writer = scanner.PostgresInventoryWriter(config)
        values = dict(zip(POSTGRES_COLUMNS, writer.row_values(self.sample_result())))

        self.assertEqual(values["owner_user_id"], "42")
        self.assertEqual(values["owner_user_login"], "alice")
        self.assertEqual(values["organization"], "FabrikamCloud")
        self.assertEqual(
            values["branch_contributing_developers"],
            "Alice Adams <alice@example.com>; Bob Brown <bob@example.com>",
        )

    def test_postgres_rows_keep_result_provider_for_mixed_scans(self):
        config = scanner.ScanConfig(
            org="FabrikamGH",
            pat="token",
            project=None,
            out_dir=Path("reports"),
            out_prefix="scan",
            max_workers=1,
            branch_workers=1,
            content_workers=1,
            max_commits_per_repo=0,
            timeout_seconds=30,
            min_confidence="low",
            provider="mixed",
        )
        writer = scanner.PostgresInventoryWriter(config)
        result = self.sample_result()
        result["provider"] = "github-enterprise"

        values = dict(zip(POSTGRES_COLUMNS, writer.row_values(result)))

        self.assertEqual(values["provider"], "github-enterprise")

    def test_category_columns_are_excel_filter_friendly(self):
        columns = scanner.category_columns(
            ["android", "react_native", "ai_enabled", "ml_enabled"]
        )

        self.assertEqual(columns["category_android"], "TRUE")
        self.assertEqual(columns["category_react_native"], "TRUE")
        self.assertEqual(columns["category_ai_enabled"], "TRUE")
        self.assertEqual(columns["category_ml_enabled"], "TRUE")
        self.assertEqual(columns["category_ios"], "FALSE")

    def test_type_columns_are_excel_filter_friendly(self):
        columns = scanner.type_columns(
            ["web_app", "microservice", "ai_enabled", "ml_enabled"]
        )

        self.assertEqual(columns["type_web_app"], "TRUE")
        self.assertEqual(columns["type_microservice"], "TRUE")
        self.assertEqual(columns["type_ai_enabled"], "TRUE")
        self.assertEqual(columns["type_ml_enabled"], "TRUE")
        self.assertEqual(columns["type_mobile_app"], "FALSE")

    def test_inventory_types_from_categories(self):
        self.assertEqual(
            scanner.inventory_types_from_categories(
                ["web_backend", "api_service", "middleware", "ml_inference"]
            ),
            ["web_app", "api_service", "middleware", "ai_enabled", "ml_enabled"],
        )

    def test_identifier_status(self):
        self.assertEqual(scanner.identifier_status("com.fabrikam.agsnap"), "found")
        self.assertEqual(scanner.identifier_status(""), "missing_from_scanned_files")

    def test_branch_name_from_ref(self):
        self.assertEqual(
            scanner.branch_name_from_ref("refs/heads/release/1.0"), "release/1.0"
        )
        self.assertEqual(scanner.branch_name_from_ref("main"), "main")

    def test_default_branch_name_from_repo(self):
        self.assertEqual(
            scanner.default_branch_name_from_repo(
                {"defaultBranch": "refs/heads/master"}
            ),
            "master",
        )
        self.assertEqual(
            scanner.default_branch_name_from_repo(
                {"defaultBranch": "refs/heads/develop"}
            ),
            "develop",
        )
        self.assertEqual(scanner.default_branch_name_from_repo({}), "")

    def test_list_branch_targets_uses_only_default_branch(self):
        client = self.FakeBranchClient()
        target = scanner.RepoScanTarget(
            project_name="Project",
            repo={
                "id": "repo-id",
                "name": "Repo",
                "defaultBranch": "refs/heads/release",
            },
        )

        branch_targets = scanner.list_branch_targets(client, target)

        self.assertEqual(len(branch_targets), 1)
        self.assertEqual(branch_targets[0].branch_name, "release")
        self.assertEqual(client.branch_calls, 0)

    def test_select_fallback_branch_name_prefers_deployment_names(self):
        self.assertEqual(
            scanner.select_fallback_branch_name(
                ["feature/foo", "main", "release/prod"]
            ),
            "release/prod",
        )
        self.assertEqual(
            scanner.select_fallback_branch_name(["feature/foo", "development", "main"]),
            "main",
        )

    def test_list_branch_targets_uses_pipeline_branch_when_default_missing(self):
        client = self.FakeBranchClient(
            refs=[
                {"name": "refs/heads/main"},
                {"name": "refs/heads/release/prod"},
                {"name": "refs/heads/development"},
            ],
            definitions=[
                {
                    "repository": {"defaultBranch": "refs/heads/release/prod"},
                    "triggers": [{"branchFilters": ["+refs/heads/release/prod"]}],
                }
            ],
        )
        target = scanner.RepoScanTarget(
            project_name="Project",
            repo={
                "id": "repo-id",
                "name": "Repo",
            },
        )

        branch_targets = scanner.list_branch_targets(client, target)

        self.assertEqual(len(branch_targets), 1)
        self.assertEqual(branch_targets[0].branch_name, "release/prod")
        self.assertEqual(client.branch_calls, 1)
        self.assertEqual(client.definition_calls, 1)

    def test_list_branch_targets_uses_keyword_branch_when_default_missing(self):
        client = self.FakeBranchClient(
            refs=[
                {"name": "refs/heads/feature/foo"},
                {"name": "refs/heads/development"},
                {"name": "refs/heads/preprod"},
            ]
        )
        target = scanner.RepoScanTarget(
            project_name="Project",
            repo={
                "id": "repo-id",
                "name": "Repo",
            },
        )

        branch_targets = scanner.list_branch_targets(client, target)

        self.assertEqual(len(branch_targets), 1)
        self.assertEqual(branch_targets[0].branch_name, "preprod")

    def test_branch_age_bucket(self):
        now = datetime(2026, 6, 21, tzinfo=timezone.utc)

        self.assertEqual(
            scanner.branch_age_bucket("2026-06-01T00:00:00Z", 90, now),
            scanner.ACTIVE_SHEET_NAME,
        )
        self.assertEqual(
            scanner.branch_age_bucket("2026-01-01T00:00:00Z", 90, now),
            scanner.OLDER_SHEET_NAME,
        )
        self.assertEqual(
            scanner.branch_age_bucket("", 90, now), scanner.OLDER_SHEET_NAME
        )


class StoreLookupTests(unittest.TestCase):
    def test_target_store_platforms_uses_native_categories(self):
        self.assertEqual(
            scanner.target_store_platforms(["ios"]), (scanner.APPLE_PLATFORM,)
        )
        self.assertEqual(
            scanner.target_store_platforms(["android"]), (scanner.GOOGLE_PLATFORM,)
        )
        self.assertEqual(
            scanner.target_store_platforms(["react_native"]),
            (scanner.APPLE_PLATFORM, scanner.GOOGLE_PLATFORM),
        )

    def test_normalize_store_countries_deduplicates_and_uppercases(self):
        self.assertEqual(
            scanner.normalize_store_countries("us, ca\nGB"), ("US", "CA", "GB")
        )
        self.assertEqual(scanner.normalize_store_countries([]), ("US",))
        with self.assertRaises(ValueError):
            scanner.normalize_store_countries(["USA"])

    def test_store_columns_aggregate_multiple_countries(self):
        columns = scanner.store_columns_from_listings(
            [
                scanner.StoreListing(
                    platform=scanner.APPLE_PLATFORM,
                    status="found",
                    name="Agsnap",
                    identifier="com.fabrikam.agsnap",
                    version="1.0.2",
                    country="US",
                ),
                scanner.StoreListing(
                    platform=scanner.APPLE_PLATFORM,
                    status="found",
                    name="Agsnap",
                    identifier="com.fabrikam.agsnap",
                    version="1.0.3",
                    country="CA",
                ),
            ]
        )

        self.assertEqual(columns["store_validation_passed"], "TRUE")
        self.assertEqual(columns["store_platforms"], "Apple App Store (US, CA)")
        self.assertEqual(columns["apple_app_store_version"], "1.0.2; 1.0.3")

    def test_store_columns_from_found_listings(self):
        columns = scanner.store_columns_from_listings(
            [
                scanner.StoreListing(
                    platform=scanner.APPLE_PLATFORM,
                    status="found",
                    name="Agsnap",
                    identifier="com.fabrikam.agsnap",
                    url="https://apps.apple.com/app/id123",
                    version="1.0.2",
                    last_updated="2026-01-01T00:00:00Z",
                ),
                scanner.StoreListing(
                    platform=scanner.GOOGLE_PLATFORM,
                    status="not_found_publicly",
                    identifier="com.fabrikam.agsnap",
                ),
            ]
        )

        self.assertEqual(columns["store_lookup_status"], "partial_found")
        self.assertEqual(columns["store_validation_passed"], "FALSE")
        self.assertEqual(columns["store_platforms"], "Apple App Store")
        self.assertEqual(columns["apple_app_store_name"], "Agsnap")
        self.assertEqual(columns["apple_app_store_identifier"], "com.fabrikam.agsnap")
        self.assertEqual(columns["apple_app_store_validation_passed"], "TRUE")
        self.assertEqual(columns["google_play_validation_passed"], "FALSE")
        self.assertEqual(columns["google_play_lookup_status"], "not_found_publicly")

    def test_store_validation_passes_when_all_requested_stores_are_found(self):
        columns = scanner.store_columns_from_listings(
            [
                scanner.StoreListing(
                    platform=scanner.APPLE_PLATFORM,
                    status="found",
                    identifier="com.fabrikam.agsnap",
                ),
                scanner.StoreListing(
                    platform=scanner.GOOGLE_PLATFORM,
                    status="not_requested",
                ),
            ]
        )

        self.assertEqual(columns["store_validation_passed"], "TRUE")
        self.assertEqual(columns["apple_app_store_validation_passed"], "TRUE")
        self.assertEqual(columns["google_play_validation_passed"], "FALSE")

    def test_store_columns_disabled_and_identifier_missing(self):
        disabled = scanner.store_columns("com.fabrikam.agsnap", ["android"], None)
        missing = scanner.store_columns("", ["ios"], object())

        self.assertEqual(disabled["store_lookup_status"], "disabled")
        self.assertEqual(disabled["store_validation_passed"], "FALSE")
        self.assertEqual(disabled["google_play_lookup_status"], "disabled")
        self.assertEqual(disabled["google_play_validation_passed"], "FALSE")
        self.assertEqual(missing["store_lookup_status"], "identifier_missing")
        self.assertEqual(missing["store_validation_passed"], "FALSE")
        self.assertEqual(missing["apple_app_store_lookup_status"], "identifier_missing")

    def test_store_columns_rejects_invalid_identifiers_without_network_lookup(self):
        class FailingStoreClient:
            def lookup(self, identifier, categories):
                raise AssertionError("invalid identifiers should not call store lookup")

        columns = scanner.store_columns(
            "#ID#Configuration:CONNECTIONS_CONFIG", ["android"], FailingStoreClient()
        )

        self.assertEqual(columns["store_lookup_status"], "identifier_invalid")
        self.assertEqual(columns["store_validation_passed"], "FALSE")
        self.assertEqual(columns["google_play_lookup_status"], "identifier_invalid")
        self.assertEqual(
            columns["google_play_identifier"], "#ID#Configuration:CONNECTIONS_CONFIG"
        )
        self.assertFalse(
            scanner.is_store_identifier_candidate(
                "#ID#Configuration:CONNECTIONS_CONFIG"
            )
        )
        self.assertTrue(scanner.is_store_identifier_candidate("com.fabrikam.agsnap"))

    def test_store_columns_reports_tls_errors_without_retrying_reads(self):
        columns = scanner.store_columns_from_listings(
            [
                scanner.StoreListing(
                    platform=scanner.GOOGLE_PLATFORM,
                    status="tls_error",
                    identifier="com.fabrikam.agsnap",
                )
            ]
        )

        self.assertEqual(columns["store_lookup_status"], "error")
        self.assertEqual(columns["store_validation_passed"], "FALSE")
        self.assertEqual(columns["google_play_lookup_status"], "tls_error")

    def test_google_play_helpers(self):
        html = """\
<html>
<head>
  <title>Agsnap - Apps on Google Play</title>
  <meta property="og:title" content="Agsnap - Apps on Google Play" />
  <meta property="og:url" content="https://play.google.com/store/apps/details?id=com.fabrikam.agsnap" />
  <script>{"softwareVersion":"1.0.2","dateModified":"2026-01-02"}</script>
</head>
</html>
"""
        parser = scanner.MetaTagParser()
        parser.feed(html)

        self.assertEqual(
            scanner.normalize_google_play_title(parser.meta["og:title"]), "Agsnap"
        )
        self.assertTrue(
            scanner.google_play_app_page(
                parser.meta, parser.meta["og:title"], "com.fabrikam.agsnap"
            )
        )
        self.assertFalse(
            scanner.google_play_app_page({}, "Google Play", "com.fabrikam.agsnap")
        )
        self.assertEqual(scanner.extract_google_play_version(html), "1.0.2")
        self.assertEqual(scanner.extract_google_play_updated(html), "2026-01-02")


class CliTests(unittest.TestCase):
    def test_parse_args_supports_multiple_github_urls_and_defaults_api_url(self):
        with patch.dict("os.environ", {"GITHUB_TOKEN": "token"}, clear=True):
            config = scanner.parse_args(
                [
                    "--provider",
                    "github-enterprise",
                    "--github-url",
                    "https://github.com/global-snt",
                    "--github-url",
                    "security-team",
                    "--out-dir",
                    "reports",
                ]
            )

        self.assertEqual(config.github_urls, ("global-snt", "security-team"))
        self.assertEqual(config.org, "global-snt")
        self.assertEqual(config.base_url, "https://api.github.com")

    def test_parse_args_requires_pat(self):
        with self.assertRaises(SystemExit):
            scanner.parse_args(["--org", "example"])

    def test_confidence_rank_rejects_unknown_values(self):
        with self.assertRaises(Exception):
            scanner.confidence_rank("banana")

    def test_parse_args_rejects_invalid_content_workers(self):
        with self.assertRaises(SystemExit):
            scanner.parse_args(
                [
                    "--org",
                    "example",
                    "--pat",
                    "token",
                    "--content-workers",
                    "0",
                ]
            )

    def test_parse_args_rejects_invalid_commit_limit(self):
        with self.assertRaises(SystemExit):
            scanner.parse_args(
                [
                    "--org",
                    "example",
                    "--pat",
                    "token",
                    "--max-commits-per-repo",
                    "-1",
                ]
            )

    def test_parse_args_rejects_invalid_branch_age_days(self):
        with self.assertRaises(SystemExit):
            scanner.parse_args(
                [
                    "--org",
                    "example",
                    "--pat",
                    "token",
                    "--branch-age-days",
                    "0",
                ]
            )

    def test_parse_args_accepts_store_lookup_options(self):
        config = scanner.parse_args(
            [
                "--org",
                "example",
                "--pat",
                "token",
                "--branch-workers",
                "24",
                "--activity-mode",
                "latest",
                "--store-lookup",
                "--store-country",
                "ca",
                "--store-country",
                "us",
                "--store-timeout",
                "7",
            ]
        )

        self.assertEqual(config.branch_workers, 24)
        self.assertEqual(config.activity_mode, "latest")
        self.assertTrue(config.store_lookup)
        self.assertEqual(config.store_country, "CA")
        self.assertEqual(config.store_countries, ("CA", "US"))
        self.assertEqual(config.store_timeout_seconds, 7)

    def test_parse_args_rejects_invalid_store_options(self):
        with self.assertRaises(SystemExit):
            scanner.parse_args(
                ["--org", "example", "--pat", "token", "--store-country", "usa"]
            )
        with self.assertRaises(SystemExit):
            scanner.parse_args(
                ["--org", "example", "--pat", "token", "--store-timeout", "0"]
            )
        with self.assertRaises(SystemExit):
            scanner.parse_args(
                [
                    "--org",
                    "example",
                    "--pat",
                    "token",
                    "--application-type",
                    "web_app",
                    "--store-lookup",
                ]
            )

    def test_parse_args_rejects_invalid_branch_workers(self):
        with self.assertRaises(SystemExit):
            scanner.parse_args(
                ["--org", "example", "--pat", "token", "--branch-workers", "0"]
            )


if __name__ == "__main__":
    unittest.main()
