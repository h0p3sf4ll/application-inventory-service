from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timezone
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", fgColor="12365A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def rows_to_csv(rows: Iterable[Mapping[str, Any]], columns: Sequence[str]) -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=list(columns), extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(
            {column: spreadsheet_cell(row.get(column)) for column in columns}
        )
    return buffer.getvalue().encode("utf-8-sig")


def rows_to_json(rows: Iterable[Mapping[str, Any]]) -> bytes:
    buffer = io.StringIO()
    buffer.write("[\n")
    for index, row in enumerate(rows):
        if index:
            buffer.write(",\n")
        buffer.write(json.dumps(dict(row), default=json_cell, indent=2))
    buffer.write("\n]\n")
    return buffer.getvalue().encode("utf-8")


def rows_to_xlsx(
    rows: Iterable[Mapping[str, Any]],
    columns: Sequence[str],
    sheet_name: str = "Inventory",
) -> bytes:
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(sheet_name[:31] or "Inventory")
    worksheet.freeze_panes = "A2"
    header: list[WriteOnlyCell] = []
    for column in columns:
        cell = WriteOnlyCell(worksheet, value=column.replace("_", " ").title())
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        header.append(cell)
    worksheet.append(header)
    for row in rows:
        worksheet.append([xlsx_cell(row.get(column)) for column in columns])
    for index, column in enumerate(columns, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = min(
            34, max(12, len(column) + 2)
        )
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def spreadsheet_cell(value: Any) -> str:
    text = text_cell(value)
    return f"'{text}" if text.startswith(FORMULA_PREFIXES) else text


def xlsx_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return (
            value.astimezone(timezone.utc).replace(tzinfo=None)
            if value.tzinfo
            else value
        )
    if isinstance(value, (bool, int, float)):
        return value
    return spreadsheet_cell(value)


def json_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def text_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, default=json_cell, sort_keys=True)
    return str(value).strip()
